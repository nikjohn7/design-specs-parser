"""Column mapping utilities for Excel schedule parsing.

This module provides functionality to map worksheet column headers to
canonical column names using exact matching and fuzzy matching.

Key functions:
- map_columns: Map header row columns to canonical names
"""

import re
from difflib import SequenceMatcher
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


# Extended header synonyms mapping canonical column names to their variations
# All variations should be lowercase for case-insensitive matching
# This extends the base HEADER_SYNONYMS from sheet_detector.py with additional variations
COLUMN_SYNONYMS: dict[str, list[str]] = {
    # Document/spec code column
    'doc_code': [
        'spec code',
        'code',
        'ref',
        'reference',
        'item code',
        'product code',
        'sku',
        'id',
        'item ref',
        'product ref',
        'item no',
        'item number',
        'product no',
        'product number',
    ],

    # Product name / item name column (common in normalized schedules)
    'product_name': [
        'product name',
        'item name',
        'product',
        'name',
        'item',
    ],
    
    # Image column
    'image': [
        'image',
        'photo',
        'indicative image',
        'picture',
        'item image',
        'img',
        'thumbnail',
        'product image',
        'finish image',  # Sample3 has this
        'feature image',
    ],
    
    # Item/location/description column
    'item_location': [
        'location',
        'description',
        'item & location',
        'item and location',
        'area',
        'room',
        'space',
        'item/location',
        'item description',
        'product description',
    ],
    
    # Specifications column
    'specs': [
        'specification',
        'specifications',
        'specs',
        'notes/comments',
        'details',
        'spec',
        'product details',
        'product specs',
        'technical specs',
        'technical specifications',
    ],
    
    # Manufacturer/supplier column
    'manufacturer': [
        'manufacturer',
        'supplier',
        'brand',
        'vendor',
        'maker',
        'manufacturer / supplier',
        'manufacturer/supplier',
        'make',
        'company',
        'manufacturer & supplier',
        'manufacturer and supplier',
        'supplier/manufacturer',
    ],
    
    # Notes/comments column
    'notes': [
        'notes',
        'comments',
        'remarks',
        'note',
        'comment',
        'additional notes',
        'additional comments',
        'notes (supplier',  # Handle "notes (supplier/fabric code)" pattern
    ],
    
    # Quantity column
    'qty': [
        'qty',
        'quantity',
        'count',
        'units',
        'no.',
        'number',
        'amount',
        'pcs',
        'pieces',
        'unit qty',
        'unit quantity',
    ],
    
    # Cost/price column (primary - unit cost)
    'cost': [
        'cost',
        'rrp',
        'price',
        'indicative cost',
        'cost per unit',
        'cost per unit $',
        '$',
        'unit price',
        'unit cost',
        'value',
        'rate',
        'unit rate',
        'each',
        'per unit',
    ],
    
    # Total cost column (separate from unit cost)
    'total_cost': [
        'total cost',
        'total cost $',
        'total price',
        'total value',
        'extended cost',
        'extended price',
        'line total',
        'subtotal',
        'sub total',
        # Note: 'total' alone is too generic and causes false positives
    ],
    
    # Finish column (for sample3 which has separate finish column)
    'finish': [
        'finish',
        'surface',
        'surface finish',
        'coating',
        'treatment',
    ],
    
    # Material column
    'material': [
        'material',
        'composition',
        'species',
        'substrate',
        'base material',
    ],
    
    # Colour column
    'colour': [
        'colour',
        'color',
        'col',
        'shade',
        'tint',
    ],
    
    # Dimensions columns (for schedules with separate dimension columns)
    'width': [
        'width',
        'w',
        'wide',
    ],
    'length': [
        'length',
        'l',
        'len',
        'long',
        'depth',
        'd',
    ],
    'height': [
        'height',
        'h',
        'ht',
        'thickness',
        'thk',
    ],
    
    # Size column (combined dimensions)
    'size': [
        'size',
        'dimensions',
        'dims',
        'dim',
        'measurements',
    ],
    
    # Lead time column
    'lead_time': [
        'lead time',
        'leadtime',
        'delivery',
        'delivery time',
        'eta',
        'availability',
    ],
    
    # Client/customer columns (for sample3)
    'client_discount': [
        'customer discount',
        'client discount',
        'discount',
        'disc',
        'disc %',
        'discount %',
    ],
    
    'client_signoff': [
        'client initials',
        'client sign off',
        'client signoff',
        'sign off',
        'signoff',
        'approval',
        'approved',
        'initials',
    ],
    
    # Trade price columns
    'trade_price': [
        'trade',
        'trade $',
        'trade price',
        'trade cost',
        'wholesale',
        'wholesale price',
    ],
}

# Minimum similarity ratio for fuzzy matching (0.0 to 1.0)
FUZZY_MATCH_THRESHOLD = 0.75


def _normalize_header(text: str | None) -> str:
    """Normalize header text for comparison.
    
    Performs the following normalizations:
    - Convert to lowercase
    - Strip leading/trailing whitespace
    - Replace newlines and multiple spaces with single space
    - Remove common suffixes like $ symbols at the end
    
    Args:
        text: Header text to normalize
        
    Returns:
        Normalized header text, or empty string if input is None/empty
    """
    if text is None:
        return ''
    
    if not isinstance(text, str):
        text = str(text)
    
    # Convert to lowercase and strip
    text = text.lower().strip()
    
    # Replace newlines and multiple whitespace with single space
    # This handles multi-line headers like "Item\nImage" -> "item image"
    text = re.sub(r'\s+', ' ', text)
    
    # Remove trailing special characters that might be formatting
    text = text.rstrip(':.-')
    
    return text


# Build a reverse lookup: normalized header text -> canonical name
_COLUMN_LOOKUP: dict[str, str] = {}
for canonical, synonyms in COLUMN_SYNONYMS.items():
    for synonym in synonyms:
        normalized_synonym = _normalize_header(synonym)
        if normalized_synonym:
            _COLUMN_LOOKUP[normalized_synonym] = canonical


def _exact_match(text: str) -> str | None:
    """Try to find an exact match for the header text.
    
    Args:
        text: Normalized header text
        
    Returns:
        Canonical column name if matched, None otherwise
    """
    if not text:
        return None
    
    # Direct lookup
    if text in _COLUMN_LOOKUP:
        return _COLUMN_LOOKUP[text]
    
    # Try partial matching for compound headers
    # e.g., "item & location (see notes)" should match "item & location"
    # But be careful not to match too broadly (e.g., "code" in "fabric code")
    
    # Sort synonyms by length (longest first) to prefer more specific matches
    sorted_synonyms = sorted(_COLUMN_LOOKUP.items(), key=lambda x: len(x[0]), reverse=True)
    
    for synonym, canonical in sorted_synonyms:
        # Skip very short synonyms for partial matching to avoid false positives
        if len(synonym) < 3:
            continue
            
        # Check if text starts with the synonym
        if text.startswith(synonym):
            # Make sure it's a word boundary (not in the middle of a word)
            if len(text) == len(synonym) or not text[len(synonym)].isalnum():
                return canonical
        
        # Check if synonym is contained in text as a complete phrase
        # Use word boundary check to avoid matching "code" in "fabric code"
        pattern = r'\b' + re.escape(synonym) + r'\b'
        if re.search(pattern, text):
            return canonical
    
    return None


def _fuzzy_match(text: str, threshold: float = FUZZY_MATCH_THRESHOLD) -> tuple[str | None, float]:
    """Try to find a fuzzy match for the header text.
    
    Uses difflib.SequenceMatcher to find the best matching synonym.
    
    Args:
        text: Normalized header text
        threshold: Minimum similarity ratio to accept a match (0.0 to 1.0)
        
    Returns:
        Tuple of (canonical column name, similarity ratio) if matched above threshold,
        (None, 0.0) otherwise
    """
    if not text:
        return None, 0.0
    
    best_match: str | None = None
    best_ratio = 0.0
    
    for synonym, canonical in _COLUMN_LOOKUP.items():
        # Calculate similarity ratio
        ratio = SequenceMatcher(None, text, synonym).ratio()
        
        if ratio > best_ratio:
            best_ratio = ratio
            best_match = canonical
    
    if best_ratio >= threshold:
        return best_match, best_ratio
    
    return None, 0.0


def _match_column(text: str, use_fuzzy: bool = True) -> tuple[str | None, str]:
    """Match header text to a canonical column name.
    
    First tries exact matching, then falls back to fuzzy matching if enabled.
    
    Args:
        text: Normalized header text
        use_fuzzy: Whether to use fuzzy matching as fallback
        
    Returns:
        Tuple of (canonical column name, match type) where match type is
        'exact', 'partial', 'fuzzy', or 'none'
    """
    if not text:
        return None, 'none'
    
    # Try exact match first
    canonical = _exact_match(text)
    if canonical:
        # Determine if it was direct or partial match
        if text in _COLUMN_LOOKUP:
            return canonical, 'exact'
        return canonical, 'partial'
    
    # Try fuzzy match if enabled
    if use_fuzzy:
        canonical, ratio = _fuzzy_match(text)
        if canonical:
            return canonical, 'fuzzy'
    
    return None, 'none'


def map_columns(
    ws: "Worksheet",
    header_row: int,
    max_cols: int = 30,
    use_fuzzy: bool = True,
) -> dict[str, int]:
    """Map worksheet column headers to canonical column names.
    
    Scans the specified header row and maps each column to a canonical
    column name using exact matching and optional fuzzy matching.
    
    Args:
        ws: openpyxl Worksheet object to examine
        header_row: Row number containing headers (1-indexed)
        max_cols: Maximum number of columns to scan (default 30)
        use_fuzzy: Whether to use fuzzy matching for unrecognized headers
        
    Returns:
        Dict mapping canonical column names to column indices (1-indexed).
        Only the first occurrence of each canonical name is included.
        
    Example:
        >>> from openpyxl import load_workbook
        >>> wb = load_workbook("schedule.xlsx")
        >>> ws = wb.active
        >>> columns = map_columns(ws, header_row=4)
        >>> print(columns)
        {'doc_code': 1, 'image': 2, 'item_location': 3, 'specs': 4, ...}
    """
    columns: dict[str, int] = {}
    
    # Determine actual column range to scan
    actual_max = min(max_cols, ws.max_column or 1)
    
    for col in range(1, actual_max + 1):
        cell = ws.cell(row=header_row, column=col)
        value = cell.value
        
        if value is None:
            continue
        
        normalized = _normalize_header(value)
        if not normalized:
            continue
        
        canonical, match_type = _match_column(normalized, use_fuzzy=use_fuzzy)
        
        if canonical and canonical not in columns:
            # Only store first occurrence of each canonical name
            columns[canonical] = col
    
    return columns


def get_column_mapping_details(
    ws: "Worksheet",
    header_row: int,
    max_cols: int = 30,
    use_fuzzy: bool = True,
) -> list[dict]:
    """Get detailed column mapping information for debugging.
    
    Similar to map_columns but returns detailed information about each
    column including the original header text, normalized text, canonical
    name, and match type.
    
    Args:
        ws: openpyxl Worksheet object to examine
        header_row: Row number containing headers (1-indexed)
        max_cols: Maximum number of columns to scan (default 30)
        use_fuzzy: Whether to use fuzzy matching for unrecognized headers
        
    Returns:
        List of dicts with column mapping details:
        [
            {
                'column': 1,
                'original': 'SPEC CODE',
                'normalized': 'spec code',
                'canonical': 'doc_code',
                'match_type': 'exact'
            },
            ...
        ]
    """
    details = []
    
    # Determine actual column range to scan
    actual_max = min(max_cols, ws.max_column or 1)
    
    for col in range(1, actual_max + 1):
        cell = ws.cell(row=header_row, column=col)
        value = cell.value
        
        original = str(value) if value is not None else None
        normalized = _normalize_header(value)
        canonical, match_type = _match_column(normalized, use_fuzzy=use_fuzzy)
        
        details.append({
            'column': col,
            'original': original,
            'normalized': normalized if normalized else None,
            'canonical': canonical,
            'match_type': match_type,
        })
    
    return details


# Convenience function to get all recognized canonical column names
def get_canonical_columns() -> list[str]:
    """Get list of all recognized canonical column names.
    
    Returns:
        List of canonical column names that can be mapped
    """
    return list(COLUMN_SYNONYMS.keys())


# Convenience function to get synonyms for a canonical column
def get_synonyms(canonical: str) -> list[str]:
    """Get list of synonyms for a canonical column name.
    
    Args:
        canonical: Canonical column name
        
    Returns:
        List of synonyms for the canonical name, or empty list if not found
    """
    return COLUMN_SYNONYMS.get(canonical, [])
