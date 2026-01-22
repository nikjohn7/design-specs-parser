"""Merged cell handling utilities.

This module provides functionality to fill merged cell regions in Excel
worksheets, ensuring that all cells within a merged range contain the
same value as the top-left cell.

This is essential for consistent cell reading during parsing, as openpyxl
only stores the value in the top-left cell of a merged range.
"""

from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


def fill_merged_regions(ws: "Worksheet") -> None:
    """Fill all merged cell regions with the top-left cell's value.
    
    In Excel, merged cells display a single value across multiple cells,
    but openpyxl only stores the value in the top-left cell of the range.
    This function unmerges cells and copies the top-left value to all cells
    in each merged region, enabling consistent cell access during parsing.
    
    This function modifies the worksheet in-place and should be called
    before any cell reading operations.
    
    Args:
        ws: openpyxl Worksheet object to process
        
    Returns:
        None (modifies worksheet in-place)
        
    Example:
        >>> from openpyxl import load_workbook
        >>> wb = load_workbook("schedule.xlsx")
        >>> ws = wb.active
        >>> fill_merged_regions(ws)
        >>> # Now all cells in merged regions have the same value
        >>> # as their top-left cell
        
    Note:
        - This function unmerges cells to allow value assignment
        - After calling this function, merged_cells.ranges will be empty
        - Formula cells are copied as-is (the formula string, not the result)
        - None values in top-left cells result in None for all cells in range
    """
    # Get a copy of merged ranges to iterate over
    # We need a copy because we'll be unmerging cells
    merged_ranges = list(ws.merged_cells.ranges)
    
    # Store the values and ranges before unmerging
    ranges_with_values = []
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_cell = ws.cell(row=min_row, column=min_col)
        value = top_left_cell.value
        ranges_with_values.append((merged_range, value, min_col, min_row, max_col, max_row))
    
    # Unmerge all cells first (this converts MergedCell objects to regular Cell objects)
    for merged_range, _, _, _, _, _ in ranges_with_values:
        try:
            ws.unmerge_cells(str(merged_range))
        except (KeyError, ValueError):
            # Handle edge cases where merged cells may not be properly initialized
            # This can happen with some malformed or synthetic Excel files
            # We'll manually remove the merged range from the set
            try:
                ws.merged_cells.remove(merged_range)
            except (KeyError, ValueError):
                pass  # Already removed or doesn't exist
    
    # Now fill the values into all cells of each former merged range
    for _, value, min_col, min_row, max_col, max_row in ranges_with_values:
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                try:
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                except AttributeError:
                    # Cell is still a MergedCell object (couldn't be unmerged properly)
                    # This can happen with malformed Excel files
                    # Skip this cell - the value will still be accessible via
                    # get_merged_cell_value() if needed
                    pass


def get_merged_cell_value(ws: "Worksheet", row: int, column: int) -> any:
    """Get the effective value of a cell, handling merged regions.
    
    If the cell is part of a merged region, returns the value from
    the top-left cell of that region. Otherwise, returns the cell's
    own value.
    
    This is an alternative to fill_merged_regions() when you only need
    to read specific cells without modifying the worksheet.
    
    Args:
        ws: openpyxl Worksheet object
        row: Row number (1-indexed)
        column: Column number (1-indexed)
        
    Returns:
        The effective value of the cell (from top-left if merged)
        
    Example:
        >>> value = get_merged_cell_value(ws, 5, 2)
        >>> # Returns the value even if cell B5 is part of a merged range
    """
    # Check if this cell is part of any merged range
    for merged_range in ws.merged_cells.ranges:
        if _cell_in_range(row, column, merged_range):
            # Get the top-left cell value
            min_col, min_row, _, _ = merged_range.bounds
            return ws.cell(row=min_row, column=min_col).value
    
    # Not in a merged range, return the cell's own value
    return ws.cell(row=row, column=column).value


def _cell_in_range(row: int, column: int, merged_range) -> bool:
    """Check if a cell is within a merged range.
    
    Args:
        row: Row number (1-indexed)
        column: Column number (1-indexed)
        merged_range: openpyxl MergedCellRange object
        
    Returns:
        True if the cell is within the merged range
    """
    min_col, min_row, max_col, max_row = merged_range.bounds
    return (min_row <= row <= max_row) and (min_col <= column <= max_col)


def is_merged_cell(ws: "Worksheet", row: int, column: int) -> bool:
    """Check if a cell is part of a merged region.
    
    Args:
        ws: openpyxl Worksheet object
        row: Row number (1-indexed)
        column: Column number (1-indexed)
        
    Returns:
        True if the cell is part of any merged region
    """
    for merged_range in ws.merged_cells.ranges:
        if _cell_in_range(row, column, merged_range):
            return True
    return False


def is_merged_cell_topleft(ws: "Worksheet", row: int, column: int) -> bool:
    """Check if a cell is the top-left cell of a merged region.
    
    This is useful for identifying section headers or other merged
    content that should be processed differently.
    
    Args:
        ws: openpyxl Worksheet object
        row: Row number (1-indexed)
        column: Column number (1-indexed)
        
    Returns:
        True if the cell is the top-left of a merged region
    """
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, _, _ = merged_range.bounds
        if row == min_row and column == min_col:
            return True
    return False


def get_merged_range_for_cell(ws: "Worksheet", row: int, column: int):
    """Get the merged range that contains a cell, if any.
    
    Args:
        ws: openpyxl Worksheet object
        row: Row number (1-indexed)
        column: Column number (1-indexed)
        
    Returns:
        MergedCellRange object if cell is merged, None otherwise
    """
    for merged_range in ws.merged_cells.ranges:
        if _cell_in_range(row, column, merged_range):
            return merged_range
    return None
