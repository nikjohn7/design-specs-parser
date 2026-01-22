"""Row extraction utilities for Excel schedule parsing.

This module provides functionality to iterate over product rows in a worksheet,
handling different layout types:
- Single-row-per-product (sample1, sample2): Each product is on one row
- Grouped rows (sample3): Product starts with an "item row" followed by detail rows

Key functions:
- iter_product_rows: Main iterator that yields product row data dictionaries
"""

import re
from typing import TYPE_CHECKING, Iterator, Any

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

from app.parser.merged_cells import (
    is_merged_cell,
    is_merged_cell_topleft,
    get_merged_range_for_cell,
)


# Keys that indicate a detail row in grouped layout (sample3 style)
# These appear in the "Description" column (column D) with value in column E
DETAIL_ROW_KEYS = {
    'maker:',
    'name:',
    'finish:',
    'size:',
    'lead time:',
    'notes:',
    'leadtime:',
    'material:',
    'colour:',
    'color:',
    'brand:',
    'supplier:',
    'manufacturer:',
    'dimensions:',
    'dim:',
    'width:',
    'height:',
    'length:',
    'depth:',
    'item:',  # This indicates start of item in grouped layout
}

# Keys that indicate an item row (start of product) in grouped layout
ITEM_ROW_KEYS = {
    'item:',
}

# Patterns that indicate a non-product row (skip these)
SKIP_ROW_PATTERNS = [
    r'^delivery$',
    r'^shipping$',
    r'^freight$',
    r'^total[s]?$',
    r'^sub\s*total$',
    r'^grand\s*total$',
    r'^gst$',
    r'^tax$',
]

# Compiled skip patterns
_SKIP_PATTERNS = [re.compile(p, re.IGNORECASE) for p in SKIP_ROW_PATTERNS]


def _get_cell_value(ws: "Worksheet", row: int, col: int) -> Any:
    """Get cell value, handling None gracefully.
    
    Args:
        ws: Worksheet object
        row: Row number (1-indexed)
        col: Column number (1-indexed)
        
    Returns:
        Cell value or None
    """
    try:
        cell = ws.cell(row=row, column=col)
        return cell.value
    except (IndexError, AttributeError):
        return None


def _normalize_text(value: Any) -> str:
    """Normalize cell value to string for comparison.
    
    Args:
        value: Cell value (any type)
        
    Returns:
        Normalized lowercase string, or empty string if None
    """
    if value is None:
        return ''
    return str(value).strip().lower()


def _is_empty_row(ws: "Worksheet", row: int, col_map: dict[str, int], max_cols: int = 20) -> bool:
    """Check if a row is empty (no meaningful data in mapped columns).
    
    Args:
        ws: Worksheet object
        row: Row number (1-indexed)
        col_map: Column mapping from map_columns()
        max_cols: Maximum columns to check
        
    Returns:
        True if row has no meaningful data
    """
    # Check all mapped columns
    for canonical, col in col_map.items():
        value = _get_cell_value(ws, row, col)
        if value is not None and str(value).strip():
            return False
    
    # Also check first few columns in case col_map is incomplete
    for col in range(1, min(max_cols, ws.max_column or 1) + 1):
        value = _get_cell_value(ws, row, col)
        if value is not None and str(value).strip():
            return False
    
    return True


def _is_section_header(
    ws: "Worksheet",
    row: int,
    col_map: dict[str, int],
    doc_code_col: int | None = None,
) -> tuple[bool, str | None]:
    """Check if a row is a section header (e.g., "FLOORING", "GLASS").
    
    Section headers are typically:
    - Originally merged cells spanning multiple columns (now filled after fill_merged_regions)
    - Have text content but no doc_code-like value
    - Same value repeated across multiple columns (sign of former merged cell)
    - No meaningful data in specs/manufacturer columns (or same value as column A)
    
    Args:
        ws: Worksheet object
        row: Row number (1-indexed)
        col_map: Column mapping from map_columns()
        doc_code_col: Column index for doc_code (default: from col_map or 1)
        
    Returns:
        Tuple of (is_section_header, section_name)
    """
    if doc_code_col is None:
        doc_code_col = col_map.get('doc_code', 1)
    
    # Get value in first column (typically where section headers appear)
    first_col_value = _get_cell_value(ws, row, 1)
    
    if first_col_value is None:
        return False, None
    
    first_col_text = str(first_col_value).strip()
    if not first_col_text:
        return False, None
    
    # After fill_merged_regions, merged cells have the same value in all cells
    # Check if multiple columns have the same value (indicates former merged cell)
    same_value_count = 0
    for col in range(1, min(8, (ws.max_column or 1) + 1)):
        col_value = _get_cell_value(ws, row, col)
        if col_value is not None and str(col_value).strip() == first_col_text:
            same_value_count += 1
    
    # If 3+ columns have the same value, it was likely a merged cell (section header)
    if same_value_count >= 3:
        return True, first_col_text
    
    # Also check for section headers that are all caps with no other meaningful data
    # These have text in column A but nothing different in other key columns
    if first_col_text.isupper() and len(first_col_text) < 50:
        specs_col = col_map.get('specs')
        manufacturer_col = col_map.get('manufacturer')
        item_location_col = col_map.get('item_location')
        
        # Get values from other columns
        specs_value = specs_col and _get_cell_value(ws, row, specs_col)
        manufacturer_value = manufacturer_col and _get_cell_value(ws, row, manufacturer_col)
        item_location_value = item_location_col and _get_cell_value(ws, row, item_location_col)
        
        # Check if other columns are empty or have the same value as column A
        specs_empty_or_same = not specs_value or str(specs_value).strip() == first_col_text
        manufacturer_empty_or_same = not manufacturer_value or str(manufacturer_value).strip() == first_col_text
        item_location_empty_or_same = not item_location_value or str(item_location_value).strip() == first_col_text
        
        # If all key columns are empty or have the same value, likely a section header
        if specs_empty_or_same and manufacturer_empty_or_same and item_location_empty_or_same:
            # Additional check: section headers don't look like doc codes
            # Doc codes typically have numbers or specific patterns like "FCA-01"
            if not re.search(r'\d', first_col_text) and not re.search(r'^[A-Z]{1,3}-', first_col_text):
                return True, first_col_text
    
    return False, None


def _is_skip_row(ws: "Worksheet", row: int, col_map: dict[str, int]) -> bool:
    """Check if a row should be skipped (delivery, totals, etc.).
    
    Args:
        ws: Worksheet object
        row: Row number (1-indexed)
        col_map: Column mapping from map_columns()
        
    Returns:
        True if row should be skipped
    """
    # Check first few columns for skip patterns
    for col in range(1, min(5, (ws.max_column or 1) + 1)):
        value = _get_cell_value(ws, row, col)
        if value is None:
            continue
        
        text = _normalize_text(value)
        for pattern in _SKIP_PATTERNS:
            if pattern.match(text):
                return True
    
    # Check image column for "DELIVERY" text (sample3 pattern)
    image_col = col_map.get('image')
    if image_col:
        image_value = _get_cell_value(ws, row, image_col)
        if image_value:
            text = _normalize_text(image_value)
            if text == 'delivery':
                return True
    
    return False


def _is_detail_row(ws: "Worksheet", row: int, col_map: dict[str, int]) -> tuple[bool, str | None, str | None]:
    """Check if a row is a detail row in grouped layout (sample3 style).
    
    Detail rows have:
    - A key like "Maker:", "Name:", "Finish:" in column D (or similar)
    - Value in the adjacent column (column E)
    - NOT "Item:" which indicates an item row
    
    Note: After fill_merged_regions, the doc_code column may have values
    propagated from merged cells, so we can't rely on empty doc_code.
    Instead, we check for detail keys in the description columns.
    
    Args:
        ws: Worksheet object
        row: Row number (1-indexed)
        col_map: Column mapping from map_columns()
        
    Returns:
        Tuple of (is_detail_row, key, value)
    """
    # In sample3, the key is typically in column D (index 4) and value in column E (index 5)
    # Check columns 3-6 for key:value pattern
    
    for col in range(3, min(7, (ws.max_column or 1) + 1)):
        value = _get_cell_value(ws, row, col)
        if value is None:
            continue
        
        text = _normalize_text(value)
        
        # Skip if this is "Item:" - that's an item row, not a detail row
        if text == 'item:':
            return False, None, None
        
        # Check if this looks like a detail key (ends with ":" and is a known key)
        # Only match known detail keys to avoid false positives
        if text in DETAIL_ROW_KEYS:
            # Get the value from the next column
            next_col_value = _get_cell_value(ws, row, col + 1)
            key = text.rstrip(':').strip()
            val = str(next_col_value).strip() if next_col_value else None
            return True, key, val
        
        # Also check for generic key:value pattern but be more strict
        # Must end with ":" and have a value in the next column
        if text.endswith(':') and len(text) < 25 and len(text) > 2:
            next_col_value = _get_cell_value(ws, row, col + 1)
            if next_col_value is not None and str(next_col_value).strip():
                key = text.rstrip(':').strip()
                val = str(next_col_value).strip()
                return True, key, val
    
    return False, None, None


def _is_item_row(ws: "Worksheet", row: int, col_map: dict[str, int]) -> bool:
    """Check if a row is an item row (start of a product in grouped layout).
    
    Item rows have:
    - A doc_code value in column A, OR
    - "Item:" in column D with a value in column E (sample3 pattern)
    - Typically have qty and/or cost values
    
    Args:
        ws: Worksheet object
        row: Row number (1-indexed)
        col_map: Column mapping from map_columns()
        
    Returns:
        True if this is an item row
    """
    doc_code_col = col_map.get('doc_code', 1)
    
    # Check for doc_code in column A
    doc_code_value = _get_cell_value(ws, row, doc_code_col)
    has_doc_code = doc_code_value is not None and str(doc_code_value).strip()
    
    # Check for "Item:" pattern in columns 3-6 (sample3 pattern)
    has_item_key = False
    for col in range(3, min(7, (ws.max_column or 1) + 1)):
        value = _get_cell_value(ws, row, col)
        if value is None:
            continue
        
        text = _normalize_text(value)
        if text == 'item:':
            # Check if there's a value in the next column
            next_col_value = _get_cell_value(ws, row, col + 1)
            if next_col_value is not None and str(next_col_value).strip():
                has_item_key = True
                break
    
    # An item row has either a doc_code OR an "Item:" key
    # But if it has "Item:" key, it's definitely an item row (grouped layout)
    return has_doc_code or has_item_key


def _detect_layout_type(ws: "Worksheet", header_row: int, col_map: dict[str, int], sample_rows: int = 50) -> str:
    """Detect the layout type of the worksheet.
    
    Grouped layout (sample3 style) is characterized by:
    - "Item:" keys in column D followed by detail rows (Maker:, Name:, etc.)
    - Multiple rows per product with detail keys like Maker:, Name:, Finish:
    
    Single-row layout (sample1/2 style) is characterized by:
    - Each product on a single row
    - Multi-line text in cells (specs, manufacturer columns)
    
    Args:
        ws: Worksheet object
        header_row: Header row number (1-indexed)
        col_map: Column mapping from map_columns()
        sample_rows: Number of rows to sample for detection
        
    Returns:
        'grouped' for sample3-style grouped rows, 'single' for single-row-per-product
    """
    # Sample rows after header to detect pattern
    detail_key_count = 0  # Count of detail keys (Maker:, Name:, etc.)
    item_key_count = 0  # Count of "Item:" keys found
    
    max_row = min(header_row + sample_rows, ws.max_row or header_row + 1)
    
    for row in range(header_row + 1, max_row + 1):
        # Check columns 3-6 for key patterns (typical location for detail keys)
        for col in range(3, min(7, (ws.max_column or 1) + 1)):
            value = _get_cell_value(ws, row, col)
            if value is None:
                continue
            
            text = _normalize_text(value)
            
            # Check for "Item:" key (indicates grouped layout)
            if text == 'item:':
                item_key_count += 1
                break
            
            # Check for detail keys (Maker:, Name:, Finish:, etc.)
            if text in DETAIL_ROW_KEYS and text != 'item:':
                detail_key_count += 1
                break
    
    # If we see "Item:" keys and detail keys, it's grouped layout
    # The key indicator is the presence of "Item:" followed by detail rows
    if item_key_count > 0 and detail_key_count > 0:
        return 'grouped'
    
    # If we see many detail keys (even without Item:), it might be grouped
    if detail_key_count >= 5:
        return 'grouped'
    
    return 'single'


def _extract_row_data(
    ws: "Worksheet",
    row: int,
    col_map: dict[str, int],
) -> dict[str, Any]:
    """Extract data from a single row based on column mapping.
    
    Args:
        ws: Worksheet object
        row: Row number (1-indexed)
        col_map: Column mapping from map_columns()
        
    Returns:
        Dictionary with canonical column names as keys and cell values
    """
    data: dict[str, Any] = {'row_num': row}
    
    for canonical, col in col_map.items():
        value = _get_cell_value(ws, row, col)
        if value is not None:
            # Convert to string and strip whitespace
            if isinstance(value, str):
                value = value.strip()
            elif isinstance(value, (int, float)):
                # Keep numeric values as-is
                pass
            else:
                value = str(value).strip()
        data[canonical] = value
    
    return data


def _extract_grouped_item_data(
    ws: "Worksheet",
    row: int,
    col_map: dict[str, int],
) -> dict[str, Any]:
    """Extract data from an item row in grouped layout (sample3 style).
    
    In grouped layout, the item row has:
    - doc_code in column A
    - Area/location in column B
    - "Item:" in description column with product name in next column
    - qty, cost, rrp in their respective columns
    
    Args:
        ws: Worksheet object
        row: Row number (1-indexed)
        col_map: Column mapping from map_columns()
        
    Returns:
        Dictionary with extracted data
    """
    data = _extract_row_data(ws, row, col_map)
    
    # For grouped layout, also extract the "Item:" value
    # Look for "Item:" pattern and extract the product name
    for col in range(1, min(10, (ws.max_column or 1) + 1)):
        value = _get_cell_value(ws, row, col)
        if value is None:
            continue
        
        text = _normalize_text(value)
        if text == 'item:':
            next_col_value = _get_cell_value(ws, row, col + 1)
            if next_col_value:
                data['item_name'] = str(next_col_value).strip()
            break
    
    return data


def iter_product_rows(
    ws: "Worksheet",
    header_row: int,
    col_map: dict[str, int],
    max_rows: int | None = None,
) -> Iterator[dict[str, Any]]:
    """Iterate over product rows in a worksheet.
    
    This is the main function for extracting product data from a worksheet.
    It handles both single-row-per-product and grouped row layouts.
    
    For single-row layout (sample1, sample2):
    - Each row with a doc_code is yielded as a product
    - Section headers are detected and propagated to subsequent rows
    
    For grouped layout (sample3):
    - Item rows start a new product
    - Detail rows (Maker:, Name:, etc.) are collected and attached
    - Product is yielded when next item row is encountered
    
    Args:
        ws: Worksheet object (should have merged cells filled first)
        header_row: Header row number (1-indexed)
        col_map: Column mapping from map_columns()
        max_rows: Maximum rows to process (None for all)
        
    Yields:
        Dictionary with product data including:
        - All mapped column values
        - 'row_num': Source row number
        - 'section': Current section context (if any)
        - 'detail_rows': List of detail row dicts (for grouped layout)
        
    Example:
        >>> from openpyxl import load_workbook
        >>> from app.parser.column_mapper import map_columns
        >>> from app.parser.merged_cells import fill_merged_regions
        >>> 
        >>> wb = load_workbook("schedule.xlsx")
        >>> ws = wb.active
        >>> fill_merged_regions(ws)
        >>> col_map = map_columns(ws, header_row=4)
        >>> 
        >>> for product in iter_product_rows(ws, header_row=4, col_map=col_map):
        ...     print(product['doc_code'], product.get('section'))
    """
    # Detect layout type
    layout_type = _detect_layout_type(ws, header_row, col_map)
    
    # Determine row range
    start_row = header_row + 1
    end_row = ws.max_row or start_row
    if max_rows:
        end_row = min(end_row, start_row + max_rows - 1)
    
    # Track current section for section header propagation
    current_section: str | None = None
    
    if layout_type == 'grouped':
        # Grouped layout: collect item + detail rows
        yield from _iter_grouped_rows(ws, start_row, end_row, col_map, current_section)
    else:
        # Single-row layout: yield each product row
        yield from _iter_single_rows(ws, start_row, end_row, col_map, current_section)


def _iter_single_rows(
    ws: "Worksheet",
    start_row: int,
    end_row: int,
    col_map: dict[str, int],
    current_section: str | None,
) -> Iterator[dict[str, Any]]:
    """Iterate over single-row-per-product layout.
    
    Args:
        ws: Worksheet object
        start_row: First data row (after header)
        end_row: Last row to process
        col_map: Column mapping
        current_section: Initial section context
        
    Yields:
        Product data dictionaries
    """
    for row in range(start_row, end_row + 1):
        # Skip empty rows
        if _is_empty_row(ws, row, col_map):
            continue
        
        # Check for section header
        is_section, section_name = _is_section_header(ws, row, col_map)
        if is_section:
            current_section = section_name
            continue
        
        # Skip delivery/total rows
        if _is_skip_row(ws, row, col_map):
            continue
        
        # Check if this is a product row (has doc_code or meaningful data)
        doc_code_col = col_map.get('doc_code', 1)
        doc_code_value = _get_cell_value(ws, row, doc_code_col)
        
        # For single-row layout, we need at least a doc_code or item_location
        item_location_col = col_map.get('item_location')
        item_location_value = item_location_col and _get_cell_value(ws, row, item_location_col)
        
        if not doc_code_value and not item_location_value:
            continue
        
        # Extract row data
        data = _extract_row_data(ws, row, col_map)
        data['section'] = current_section
        data['detail_rows'] = []
        
        yield data


def _has_item_key(ws: "Worksheet", row: int) -> tuple[bool, str | None]:
    """Check if a row has an "Item:" key in columns 3-6.
    
    Args:
        ws: Worksheet object
        row: Row number (1-indexed)
        
    Returns:
        Tuple of (has_item_key, item_value)
    """
    for col in range(3, min(7, (ws.max_column or 1) + 1)):
        value = _get_cell_value(ws, row, col)
        if value is None:
            continue
        
        text = _normalize_text(value)
        if text == 'item:':
            # Get the value from the next column
            next_col_value = _get_cell_value(ws, row, col + 1)
            item_value = str(next_col_value).strip() if next_col_value else None
            return True, item_value
    
    return False, None


def _iter_grouped_rows(
    ws: "Worksheet",
    start_row: int,
    end_row: int,
    col_map: dict[str, int],
    current_section: str | None,
) -> Iterator[dict[str, Any]]:
    """Iterate over grouped-row layout (sample3 style).
    
    In this layout:
    - Item rows have "Item:" in column D with product name in column E
    - Detail rows have keys like "Maker:", "Name:", "Finish:" in column D
    - Product is yielded when next item row or end is reached
    
    Note: After fill_merged_regions, doc_code values are propagated to all rows
    in a merged range, so we can't rely on empty doc_code to detect detail rows.
    Instead, we check for "Item:" key to identify item rows.
    
    Args:
        ws: Worksheet object
        start_row: First data row (after header)
        end_row: Last row to process
        col_map: Column mapping
        current_section: Initial section context
        
    Yields:
        Product data dictionaries with detail_rows attached
    """
    current_product: dict[str, Any] | None = None
    
    for row in range(start_row, end_row + 1):
        # Skip empty rows
        if _is_empty_row(ws, row, col_map):
            continue
        
        # Check for section header
        is_section, section_name = _is_section_header(ws, row, col_map)
        if is_section:
            # Yield current product before section change
            if current_product:
                yield current_product
                current_product = None
            current_section = section_name
            continue
        
        # Skip delivery/total rows
        if _is_skip_row(ws, row, col_map):
            continue
        
        # Check if this row has an "Item:" key (start of new product)
        has_item, item_value = _has_item_key(ws, row)
        if has_item:
            # Yield previous product
            if current_product:
                yield current_product
            
            # Start new product
            current_product = _extract_grouped_item_data(ws, row, col_map)
            current_product['section'] = current_section
            current_product['detail_rows'] = []
            current_product['item_name'] = item_value
            continue
        
        # Check if this is a detail row (has detail key like Maker:, Name:, etc.)
        is_detail, detail_key, detail_value = _is_detail_row(ws, row, col_map)
        if is_detail:
            if current_product:
                # Add detail to current product
                current_product['detail_rows'].append({
                    'row_num': row,
                    'key': detail_key,
                    'value': detail_value,
                })
            continue
        
        # If we get here, it's an unrecognized row type
        # Skip it - don't add random data as details
    
    # Yield final product
    if current_product:
        yield current_product


def get_product_count(
    ws: "Worksheet",
    header_row: int,
    col_map: dict[str, int],
) -> int:
    """Get the count of products in a worksheet without full extraction.
    
    This is a quick way to estimate the number of products for progress
    reporting or validation.
    
    Args:
        ws: Worksheet object
        header_row: Header row number (1-indexed)
        col_map: Column mapping from map_columns()
        
    Returns:
        Estimated number of products
    """
    count = 0
    for _ in iter_product_rows(ws, header_row, col_map):
        count += 1
    return count


def extract_all_products(
    ws: "Worksheet",
    header_row: int,
    col_map: dict[str, int],
) -> list[dict[str, Any]]:
    """Extract all products from a worksheet as a list.
    
    Convenience function that collects all products from iter_product_rows.
    
    Args:
        ws: Worksheet object
        header_row: Header row number (1-indexed)
        col_map: Column mapping from map_columns()
        
    Returns:
        List of product data dictionaries
    """
    return list(iter_product_rows(ws, header_row, col_map))
