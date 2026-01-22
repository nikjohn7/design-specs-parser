"""Sheet detection utilities for Excel schedule parsing.

This module provides functionality to detect schedule sheets within Excel
workbooks and find header rows by matching against known header synonyms.

Key functions:
- find_header_row: Locate the header row in a worksheet
- is_schedule_sheet: Determine if a worksheet contains schedule data
"""

import re
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


# Header synonyms mapping canonical column names to their variations
# All variations should be lowercase for case-insensitive matching
HEADER_SYNONYMS: dict[str, list[str]] = {
    # Document/spec code column - REQUIRED for schedule detection
    'doc_code': [
        'spec code',
        'code',
        'ref',
        'reference',
        'item code',
        'product code',
        'sku',
        'id',
    ],
    
    # Image column
    'image': [
        'image',
        'photo',
        'indicative image',
        'picture',
        'item image',
        'img',
        'thumbnail',
    ],
    
    # Item/location/description column
    'item_location': [
        'location',
        'description',
        'item & location',
        'item and location',
        'area',
        'room',
        'space',
    ],

    # Product name / item name column (common in normalized schedules)
    'product_name': [
        'product name',
        'item name',
    ],
    
    # Specifications column
    'specs': [
        'specification',
        'specifications',
        'specs',
        'notes/comments',
        'details',
        'spec',
    ],
    
    # Manufacturer/supplier column
    'manufacturer': [
        'manufacturer',
        'supplier',
        'brand',
        'vendor',
        'maker',
        'manufacturer / supplier',
        'manufacturer/supplier',
        'make',
        'company',
    ],
    
    # Notes/comments column
    'notes': [
        'notes',
        'comments',
        'remarks',
        'note',
        'comment',
    ],
    
    # Quantity column
    'qty': [
        'qty',
        'quantity',
        'count',
        'units',
        'no.',
        'number',
    ],
    
    # Cost/price column
    'cost': [
        'cost',
        'rrp',
        'price',
        'indicative cost',
        'cost per unit',
        'total cost',
        '$',
        'unit price',
        'unit cost',
        'amount',
        'value',
    ],
}

# Minimum number of recognized headers to consider a row as a header row
MIN_HEADER_MATCHES = 2

# Key columns that must be present for a sheet to be considered a schedule
# At minimum, we need a doc_code column
REQUIRED_COLUMNS = {'doc_code'}

# Additional columns that strengthen schedule detection
# Having at least one of these along with doc_code confirms it's a schedule
SUPPORTING_COLUMNS = {'item_location', 'product_name', 'specs', 'manufacturer', 'cost', 'qty'}


def _normalize_header(text: str | None) -> str:
    """Normalize header text for comparison.
    
    Performs the following normalizations:
    - Convert to lowercase
    - Strip leading/trailing whitespace
    - Replace multiple spaces with single space
    - Remove common suffixes like $ symbols at the end
    - Handle newlines (take first line only)
    
    Args:
        text: Header text to normalize
        
    Returns:
        Normalized header text, or empty string if input is None/empty
    """
    if text is None:
        return ''
    
    if not isinstance(text, str):
        text = str(text)
    
    # Take first line only (some headers have multi-line content)
    text = text.split('\n')[0]
    
    # Convert to lowercase and strip
    text = text.lower().strip()
    
    # Replace multiple spaces with single space
    text = re.sub(r'\s+', ' ', text)
    
    # Remove trailing special characters that might be formatting
    text = text.rstrip(':.-')
    
    return text


# Build a reverse lookup: normalized header text -> canonical name
_HEADER_LOOKUP: dict[str, str] = {}
for canonical, synonyms in HEADER_SYNONYMS.items():
    for synonym in synonyms:
        normalized_synonym = _normalize_header(synonym)
        if normalized_synonym:
            _HEADER_LOOKUP[normalized_synonym] = canonical


def _match_header(text: str) -> str | None:
    """Match normalized header text to a canonical column name.
    
    Args:
        text: Normalized header text
        
    Returns:
        Canonical column name if matched, None otherwise
    """
    if not text:
        return None
    
    # Direct lookup
    if text in _HEADER_LOOKUP:
        return _HEADER_LOOKUP[text]
    
    # Try partial matching for compound headers
    # e.g., "item & location (see notes)" should match "item & location"
    for synonym, canonical in _HEADER_LOOKUP.items():
        if text.startswith(synonym) or synonym in text:
            return canonical
    
    return None


def _score_row_as_header(ws: "Worksheet", row: int, max_cols: int = 20) -> tuple[int, set[str]]:
    """Score a row based on how many header synonyms it matches.
    
    Args:
        ws: Worksheet to examine
        row: Row number (1-indexed)
        max_cols: Maximum number of columns to check
        
    Returns:
        Tuple of (score, set of matched canonical column names)
    """
    matched_columns: set[str] = set()
    
    for col in range(1, max_cols + 1):
        cell = ws.cell(row=row, column=col)
        value = cell.value
        
        if value is None:
            continue
        
        normalized = _normalize_header(value)
        if not normalized:
            continue
        
        canonical = _match_header(normalized)
        if canonical:
            matched_columns.add(canonical)
    
    return len(matched_columns), matched_columns


def find_header_row(ws: "Worksheet", max_scan: int = 50) -> int | None:
    """Find the header row in a worksheet.
    
    Scans the first `max_scan` rows and scores each row based on how many
    cells match known header synonyms. Returns the row with the highest
    score if it meets the minimum threshold.
    
    Args:
        ws: openpyxl Worksheet object to examine
        max_scan: Maximum number of rows to scan (default 50)
        
    Returns:
        Row number (1-indexed) of the header row, or None if not found
        
    Example:
        >>> from openpyxl import load_workbook
        >>> wb = load_workbook("schedule.xlsx")
        >>> ws = wb.active
        >>> header_row = find_header_row(ws)
        >>> if header_row:
        ...     print(f"Header found at row {header_row}")
    """
    best_row: int | None = None
    best_score = 0
    best_columns: set[str] = set()
    
    # Determine actual row range to scan
    actual_max = min(max_scan, ws.max_row or 1)
    
    for row in range(1, actual_max + 1):
        score, matched_columns = _score_row_as_header(ws, row)
        
        # Must have at least minimum matches
        if score < MIN_HEADER_MATCHES:
            continue
        
        # Prefer rows with required columns
        has_required = bool(matched_columns & REQUIRED_COLUMNS)
        has_supporting = bool(matched_columns & SUPPORTING_COLUMNS)
        
        # Calculate weighted score
        # Required columns are worth more
        weighted_score = score
        if has_required:
            weighted_score += 2
        if has_supporting:
            weighted_score += 1
        
        # Update best if this row is better
        if weighted_score > best_score:
            best_score = weighted_score
            best_row = row
            best_columns = matched_columns
        elif weighted_score == best_score and has_required and 'doc_code' not in best_columns:
            # Prefer row with doc_code if scores are equal
            best_row = row
            best_columns = matched_columns
    
    # Final validation: must have at least required columns or strong supporting evidence
    if best_row is not None:
        has_required = bool(best_columns & REQUIRED_COLUMNS)
        has_multiple_supporting = len(best_columns & SUPPORTING_COLUMNS) >= 2
        
        if has_required or has_multiple_supporting:
            return best_row
    
    return None


def get_header_columns(ws: "Worksheet", header_row: int, max_cols: int = 20) -> dict[str, int]:
    """Get mapping of canonical column names to column indices for a header row.
    
    Args:
        ws: Worksheet to examine
        header_row: Row number containing headers (1-indexed)
        max_cols: Maximum number of columns to check
        
    Returns:
        Dict mapping canonical column names to column indices (1-indexed)
        
    Example:
        >>> columns = get_header_columns(ws, 4)
        >>> print(columns)
        {'doc_code': 1, 'image': 2, 'item_location': 3, ...}
    """
    columns: dict[str, int] = {}
    
    for col in range(1, max_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        value = cell.value
        
        if value is None:
            continue
        
        normalized = _normalize_header(value)
        if not normalized:
            continue
        
        canonical = _match_header(normalized)
        if canonical and canonical not in columns:
            # Only store first occurrence of each canonical name
            columns[canonical] = col
    
    return columns


def is_schedule_sheet(ws: "Worksheet", max_scan: int = 50) -> bool:
    """Determine if a worksheet contains schedule data.
    
    A worksheet is considered a schedule sheet if:
    1. A header row can be found with recognized column names
    2. The header row contains at least a doc_code column
    3. The header row contains at least one supporting column
    
    This helps filter out cover sheets, legend sheets, and other
    non-data sheets that may be present in the workbook.
    
    Args:
        ws: openpyxl Worksheet object to examine
        max_scan: Maximum number of rows to scan for header (default 50)
        
    Returns:
        True if the worksheet appears to contain schedule data
        
    Example:
        >>> from openpyxl import load_workbook
        >>> wb = load_workbook("schedule.xlsx")
        >>> for sheet_name in wb.sheetnames:
        ...     ws = wb[sheet_name]
        ...     if is_schedule_sheet(ws):
        ...         print(f"{sheet_name} is a schedule sheet")
    """
    # Try to find header row
    header_row = find_header_row(ws, max_scan)
    
    if header_row is None:
        return False
    
    # Get the columns found in the header
    columns = get_header_columns(ws, header_row)
    
    # Must have doc_code column
    if 'doc_code' not in columns:
        return False
    
    # Must have at least one supporting column
    has_supporting = bool(set(columns.keys()) & SUPPORTING_COLUMNS)
    
    return has_supporting


def get_schedule_sheets(wb) -> list[tuple[str, "Worksheet", int]]:
    """Get all schedule sheets from a workbook with their header rows.
    
    Args:
        wb: openpyxl Workbook object
        
    Returns:
        List of tuples: (sheet_name, worksheet, header_row)
        
    Example:
        >>> from openpyxl import load_workbook
        >>> wb = load_workbook("schedule.xlsx")
        >>> for name, ws, header_row in get_schedule_sheets(wb):
        ...     print(f"Sheet '{name}' has header at row {header_row}")
    """
    schedule_sheets = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = find_header_row(ws)
        
        if header_row is not None:
            columns = get_header_columns(ws, header_row)
            
            # Validate it's a schedule sheet
            if 'doc_code' in columns and bool(set(columns.keys()) & SUPPORTING_COLUMNS):
                schedule_sheets.append((sheet_name, ws, header_row))
    
    return schedule_sheets
