"""Workbook loading and validation utilities.

This module provides safe loading of Excel workbooks with proper
error handling for invalid, corrupt, or unsupported files.
"""

import io
import zipfile
from typing import TYPE_CHECKING, Any

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException

if TYPE_CHECKING:
    from app.core.models import Product, ParseResponse

import re
from openpyxl.cell.cell import Cell


class WorkbookLoadError(Exception):
    """Exception raised when a workbook cannot be loaded.

    This exception is raised for various loading failures including:
    - Empty file data
    - Corrupt or invalid ZIP structure
    - Invalid Excel file format
    - Password-protected files
    - Other unexpected errors during loading

    Attributes:
        message: Human-readable error description
        detail: Additional technical details (optional)
    """

    def __init__(self, message: str, detail: str | None = None):
        """Initialize WorkbookLoadError.

        Args:
            message: Brief error message describing what went wrong
            detail: Additional error details or technical information
        """
        self.message = message
        self.detail = detail
        super().__init__(message)

    def __str__(self) -> str:
        """Return string representation of the error."""
        if self.detail:
            return f"{self.message}: {self.detail}"
        return self.message


def load_workbook_safe(file_bytes: bytes) -> Workbook:
    """Safely load an Excel workbook from raw bytes.

    This function validates and loads an Excel (.xlsx) file from bytes,
    handling various error conditions gracefully.

    Args:
        file_bytes: Raw bytes of the Excel file

    Returns:
        Workbook: Loaded openpyxl Workbook object

    Raises:
        WorkbookLoadError: If the file cannot be loaded due to:
            - Empty file data
            - Corrupt or invalid ZIP structure (xlsx files are ZIP archives)
            - Invalid Excel file format
            - Password-protected files
            - Other unexpected errors

    Example:
        >>> with open("schedule.xlsx", "rb") as f:
        ...     file_bytes = f.read()
        >>> wb = load_workbook_safe(file_bytes)
        >>> print(wb.sheetnames)
        ['Sheet1', 'Sheet2']
    """
    # Validate input is not empty
    if not file_bytes:
        raise WorkbookLoadError(
            message="Empty file",
            detail="The uploaded file contains no data"
        )

    # Check minimum file size (ZIP files need at least a few bytes for header)
    # A valid xlsx file should be at least ~100 bytes (empty workbook)
    if len(file_bytes) < 100:
        raise WorkbookLoadError(
            message="Invalid file",
            detail=f"File too small ({len(file_bytes)} bytes) to be a valid Excel workbook"
        )

    # Create a file-like object from bytes
    file_stream = io.BytesIO(file_bytes)

    try:
        # Load workbook with openpyxl
        # data_only=False preserves formulas (needed for schedule name extraction)
        # read_only=False allows full access to merged cells and other features
        workbook = openpyxl.load_workbook(
            file_stream,
            data_only=False,
            read_only=False,
        )
        return workbook

    except zipfile.BadZipFile as e:
        # xlsx files are ZIP archives - this error means corrupt or not xlsx
        raise WorkbookLoadError(
            message="Invalid file format",
            detail="File is not a valid Excel workbook (corrupt or not .xlsx format)"
        ) from e

    except InvalidFileException as e:
        # openpyxl-specific validation failure
        error_str = str(e).lower()

        # Check for password protection
        if "password" in error_str or "encrypted" in error_str:
            raise WorkbookLoadError(
                message="Password-protected file",
                detail="Cannot open password-protected Excel files"
            ) from e

        # Generic invalid file
        raise WorkbookLoadError(
            message="Invalid Excel file",
            detail=str(e)
        ) from e

    except PermissionError as e:
        # Unlikely with BytesIO but handle defensively
        raise WorkbookLoadError(
            message="Permission error",
            detail="Unable to read file data"
        ) from e

    except MemoryError as e:
        # File too large to load into memory
        raise WorkbookLoadError(
            message="File too large",
            detail="The file is too large to process"
        ) from e

    except KeyError as e:
        # Missing required Excel file components (e.g., [Content_Types].xml)
        # This happens when the file is a valid ZIP but not an Excel file
        raise WorkbookLoadError(
            message="Invalid Excel file",
            detail="File is a valid ZIP archive but not a valid Excel workbook (missing required components)"
        ) from e

    except Exception as e:
        # Catch-all for unexpected errors
        # Log the actual exception type for debugging
        error_type = type(e).__name__
        raise WorkbookLoadError(
            message="Failed to load workbook",
            detail=f"Unexpected error ({error_type}): {str(e)}"
        ) from e


# Patterns that indicate a cell is NOT a schedule title (metadata labels)
METADATA_LABELS = frozenset({
    'job no.', 'job no', 'job name.', 'job name', 'revision date', 'revision no.',
    'revision no', 'revision', 'date', 'client name:', 'client name', 'project address:',
    'project address', 'version:', 'version', 'issue date:', 'issue date',
    'project:', 'project', 'rev', 'rev.', 'notes', 'legend'
})

# Pattern to detect formula references (e.g., ='[1]Cover Sheet'!A6)
FORMULA_PATTERN = re.compile(
    r"^=\s*'?\[?\d*\]?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)$",
    re.IGNORECASE
)

# Pattern to detect Cover Sheet formula specifically
COVER_SHEET_FORMULA_PATTERN = re.compile(
    r"^=\s*'?\[?\d*\]?Cover\s*Sheet'?!\$?([A-Z]+)\$?(\d+)$",
    re.IGNORECASE
)

_TRAILING_QUALIFIER_PATTERN = re.compile(r"\s*\((?:ff&e|ffe)\s*tracker\)\s*$", re.IGNORECASE)


def _clean_schedule_name(text: str) -> str:
    """Normalize schedule name strings extracted from workbook cells."""
    cleaned = str(text).strip()
    cleaned = _TRAILING_QUALIFIER_PATTERN.sub("", cleaned).strip()
    return cleaned


def _get_cell_string_value(cell: Cell) -> str | None:
    """Extract string value from a cell, handling various types.
    
    Args:
        cell: openpyxl Cell object
        
    Returns:
        String value or None if cell is empty or non-string
    """
    value = cell.value
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    # Convert numbers to string (but these are unlikely to be titles)
    return str(value).strip()


def _is_metadata_label(text: str) -> bool:
    """Check if text appears to be a metadata label rather than a title.
    
    Args:
        text: Cell text to check
        
    Returns:
        True if text looks like a metadata label
    """
    if not text:
        return True
    
    normalized = text.lower().strip()
    
    # Empty after stripping whitespace
    if not normalized:
        return True
    
    # Check against known metadata labels
    if normalized in METADATA_LABELS:
        return True
    
    # Check if it ends with common label suffixes
    if normalized.endswith(':') and len(normalized) < 30:
        return True
    
    # "SCHEDULE NAME" is a label, not a title
    if normalized == 'schedule name':
        return True
    
    return False


def _is_likely_title(text: str) -> bool:
    """Check if text appears to be a schedule title.
    
    A title typically:
    - Is longer than a few characters
    - Contains alphanumeric content
    - Is not a metadata label
    - May contain project codes, names, or "schedule" keyword
    
    Args:
        text: Cell text to check
        
    Returns:
        True if text looks like a schedule title
    """
    if not text or len(text) < 3:
        return False
    
    if _is_metadata_label(text):
        return False
    
    # Check for formula (starts with =)
    if text.startswith('='):
        return False
    
    # Check for error values
    if text.startswith('#') and text.endswith('!'):
        return False
    
    # Titles often contain these patterns
    text_lower = text.lower()
    title_indicators = ['schedule', 'project', 'interior', 'finish', 'ff&e', 'ffe']

    # If it contains a title indicator, it's likely a title
    if any(indicator in text_lower for indicator in title_indicators):
        return True

    # Disclaimers and instruction blocks are not titles (common above headers).
    if "\n" in text:
        return False
    disclaimer_indicators = [
        "refer to drawings",
        "refer to plans",
        "for full detail",
        "images and costs",
        "verify on site",
        "prior to order",
        "indicative only",
        "all dimensions",
    ]
    if any(indicator in text_lower for indicator in disclaimer_indicators):
        return False

    # Avoid mistaking column headers (e.g., "Indicative Image") for titles.
    # If it's a short phrase with no title indicators, no digits, and no colon,
    # treat it as a header/label rather than a schedule name.
    if ":" not in text and not re.search(r"\d", text):
        words = re.findall(r"[A-Za-z&]+", text)
        if 1 <= len(words) <= 3 and len(text) <= 28:
            return False
    
    # If it contains a colon followed by text (like "12006: GEM, WATERLINE PLACE")
    if ':' in text and len(text) > 10:
        parts = text.split(':', 1)
        if len(parts) == 2 and len(parts[1].strip()) > 3:
            return True
    
    # If it's reasonably long and not a label, consider it a title
    if len(text) > 15 and not text.endswith(':'):
        return True
    
    return False


def _resolve_cover_sheet_formula(wb: Workbook, formula: str) -> str | None:
    """Attempt to resolve a formula reference to Cover Sheet.
    
    Args:
        wb: Workbook object
        formula: Formula string (e.g., "='[1]Cover Sheet'!A6")
        
    Returns:
        Resolved cell value or None if cannot resolve
    """
    match = COVER_SHEET_FORMULA_PATTERN.match(formula)
    if not match:
        # Try generic formula pattern
        match = FORMULA_PATTERN.match(formula)
        if not match:
            return None
        sheet_name = match.group(1).strip()
    else:
        sheet_name = 'Cover Sheet'
    
    col_letter = match.group(1) if COVER_SHEET_FORMULA_PATTERN.match(formula) else match.group(2)
    row_num = int(match.group(2) if COVER_SHEET_FORMULA_PATTERN.match(formula) else match.group(3))
    
    # Find the Cover Sheet (case-insensitive, handle trailing spaces)
    cover_sheet = None
    for ws_name in wb.sheetnames:
        if ws_name.strip().lower() == sheet_name.lower():
            cover_sheet = wb[ws_name]
            break
    
    if cover_sheet is None:
        return None
    
    # Get the referenced cell value
    try:
        cell = cover_sheet[f"{col_letter}{row_num}"]
        value = _get_cell_string_value(cell)
        if value and _is_likely_title(value):
            return _clean_schedule_name(value)
    except (KeyError, ValueError):
        pass
    
    return None


def _find_schedule_name_in_cover_sheet(wb: Workbook) -> str | None:
    """Search for schedule name in a Cover Sheet.
    
    Looks for patterns like:
    - Row with "SCHEDULE NAME" label and value in adjacent cell
    - Title in A6 (common location)
    - PROJECT: prefix rows
    
    Args:
        wb: Workbook object
        
    Returns:
        Schedule name or None if not found
    """
    # Find Cover Sheet
    cover_sheet = None
    for ws_name in wb.sheetnames:
        if 'cover' in ws_name.strip().lower():
            cover_sheet = wb[ws_name]
            break
    
    if cover_sheet is None:
        return None
    
    # Scan first 10 rows for schedule name patterns
    for row_idx in range(1, 11):
        # Check for "SCHEDULE NAME" label pattern
        cell_a = cover_sheet.cell(row=row_idx, column=1)
        val_a = _get_cell_string_value(cell_a)
        
        if val_a:
            val_a_lower = val_a.lower().strip()
            
            # Pattern: "SCHEDULE NAME" in A, actual name in B
            if 'schedule name' in val_a_lower or val_a_lower == 'schedule':
                cell_b = cover_sheet.cell(row=row_idx, column=2)
                val_b = _get_cell_string_value(cell_b)
                if val_b and len(val_b) > 2:
                    return _clean_schedule_name(val_b)
            
            # Pattern: Title directly in A (like "SCHEDULE 003- INTERNAL FINISHES")
            if _is_likely_title(val_a) and not val_a_lower.startswith("project:"):
                return _clean_schedule_name(val_a)
    
    # Check A6 specifically (common location in sample2)
    try:
        cell_a6 = cover_sheet.cell(row=6, column=1)
        val_a6 = _get_cell_string_value(cell_a6)
        if val_a6 and _is_likely_title(val_a6):
            return _clean_schedule_name(val_a6)
    except (KeyError, ValueError):
        pass
    
    return None


def get_schedule_name(wb: Workbook, filename: str) -> str:
    """Extract the schedule name from a workbook.
    
    This function attempts to find the schedule name using multiple strategies:
    1. Check rows 1-3 of the first sheet for title text
    2. Handle formula references (e.g., ='[1]Cover Sheet'!A6)
    3. Search Cover Sheet for schedule name patterns
    4. Fallback to filename without extension
    
    Args:
        wb: Loaded openpyxl Workbook object
        filename: Original filename (used as fallback)
        
    Returns:
        Schedule name string (never empty)
        
    Example:
        >>> wb = load_workbook_safe(file_bytes)
        >>> name = get_schedule_name(wb, "schedule_sample1.xlsx")
        >>> print(name)
        '12006: GEM, WATERLINE PLACE, WILLIAMSTOWN'
    """
    if not wb.sheetnames:
        # Empty workbook, use filename
        return _filename_to_schedule_name(filename)
    
    # Get the first sheet (or active sheet)
    first_sheet = wb.active or wb[wb.sheetnames[0]]
    
    # Strategy 1: Check rows 1-10 of first sheet for title text
    project_title_candidate: str | None = None
    for row_idx in range(1, 11):
        # Check column A first
        cell_a = first_sheet.cell(row=row_idx, column=1)
        val_a = _get_cell_string_value(cell_a)
        
        if val_a:
            # Check if it's a formula reference
            if val_a.startswith('='):
                # Try to resolve formula
                resolved = _resolve_cover_sheet_formula(wb, val_a)
                if resolved:
                    return _clean_schedule_name(resolved)
                # If formula can't be resolved, continue searching
                continue
            
            # Check if it's a likely title
            if _is_likely_title(val_a):
                # Many workbooks (including synthetic ones) put a "PROJECT:" line above
                # a separate "SCHEDULE NAME" field. Prefer the schedule name field if present.
                if val_a.lower().lstrip().startswith("project:"):
                    project_title_candidate = project_title_candidate or val_a
                    continue
                return _clean_schedule_name(val_a)
            
            # Check for "SCHEDULE NAME" label pattern
            val_a_lower = val_a.lower().strip()
            if 'schedule name' in val_a_lower or val_a_lower == 'schedule':
                cell_b = first_sheet.cell(row=row_idx, column=2)
                val_b = _get_cell_string_value(cell_b)
                if val_b and len(val_b) > 2:
                    return _clean_schedule_name(val_b)
        
        # Also check column B for titles (some formats put title there)
        cell_b = first_sheet.cell(row=row_idx, column=2)
        val_b = _get_cell_string_value(cell_b)
        if val_b and _is_likely_title(val_b):
            # Make sure column A isn't a label
            if not val_a or not _is_metadata_label(val_a):
                return _clean_schedule_name(val_b)

    if project_title_candidate:
        return _clean_schedule_name(project_title_candidate)
    
    # Strategy 2: Search Cover Sheet if present
    cover_name = _find_schedule_name_in_cover_sheet(wb)
    if cover_name:
        return _clean_schedule_name(cover_name)
    
    # Strategy 3: Fallback to filename
    return _filename_to_schedule_name(filename)


def _filename_to_schedule_name(filename: str) -> str:
    """Convert filename to a schedule name.
    
    Removes file extension and cleans up the name.
    
    Args:
        filename: Original filename
        
    Returns:
        Cleaned schedule name
    """
    if not filename:
        return "Unknown Schedule"
    
    # Remove common extensions
    name = filename
    for ext in ['.xlsx', '.xls', '.XLSX', '.XLS']:
        if name.endswith(ext):
            name = name[:-len(ext)]
            break
    
    # Clean up underscores and extra spaces
    name = name.replace('_', ' ').strip()
    
    # If empty after cleaning, return default
    if not name:
        return "Unknown Schedule"
    
    return name


def _looks_like_repeated_header_row(row_data: dict[str, Any]) -> bool:
    """Heuristic to skip header rows repeated mid-sheet.

    Synthetic and real-world schedules sometimes repeat the header row in the
    middle of a sheet (e.g., after a page break). The row extractor treats these
    as potential product rows, so we filter them out here.
    """

    def norm(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip().lower()

    doc_code = norm(row_data.get("doc_code"))
    if not doc_code:
        return False

    # Must look like a header cell for the doc_code column, plus at least 2 other
    # header-like cells in common columns.
    doc_code_headers = {
        "spec code",
        "doc code",
        "drawing code",
        "code",
        "ref",
        "ref no",
        "reference",
        "id",
        "sku",
        "item code",
        "product code",
    }
    if doc_code not in doc_code_headers:
        return False

    headerish = 0
    candidates: dict[str, set[str]] = {
        "item_location": {"item & location", "item and location", "area", "room", "location", "description"},
        "specs": {"specification", "specifications", "specs", "notes/comments", "details", "spec"},
        "manufacturer": {"manufacturer", "supplier", "brand", "vendor", "maker", "manufacturer / supplier"},
        "notes": {"notes", "comments", "remarks"},
        "qty": {"qty", "quantity", "units", "no.", "no"},
        "cost": {"cost", "rrp", "price", "indicative cost", "cost per unit", "unit price", "unit cost", "$"},
    }
    for key, values in candidates.items():
        if norm(row_data.get(key)) in values:
            headerish += 1

    return headerish >= 2


def _normalize_doc_code_for_dedup(doc_code: str | None) -> str | None:
    if doc_code is None:
        return None
    normalized = doc_code.strip()
    return normalized or None


def _dedupe_products_by_doc_code(products: list["Product"]) -> list["Product"]:
    """De-duplicate products by doc_code only.

    Rules:
      - Use `doc_code` as the only key (after stripping surrounding whitespace).
      - Keep the first occurrence when duplicates are found.
      - Keep all products where `doc_code` is None/empty/whitespace.
    """

    seen: set[str] = set()
    deduped: list["Product"] = []

    for product in products:
        doc_code_key = _normalize_doc_code_for_dedup(product.doc_code)
        if doc_code_key is None:
            deduped.append(product)
            continue
        if doc_code_key in seen:
            continue
        seen.add(doc_code_key)
        deduped.append(product)

    return deduped


def parse_workbook(wb: Workbook, filename: str, extract_images: bool = False) -> "ParseResponse":
    """Parse an openpyxl workbook into a structured API response.

    This is a backward-compatible wrapper around ScheduleParser.
    For new code, prefer using ScheduleParser directly with configuration.

    Args:
        wb: Loaded openpyxl workbook.
        filename: Original filename (used for schedule_name fallback).
        extract_images: Reserved for future use (feature_image extraction is not implemented).

    Returns:
        ParseResponse with schedule_name and combined products from all schedule sheets.
    """
    from app.parser.service import ScheduleParser, ScheduleParserConfig

    config = ScheduleParserConfig(extract_images=extract_images)
    parser = ScheduleParser(config=config)
    return parser.parse_workbook(wb, filename)
