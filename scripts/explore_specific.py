#!/usr/bin/env python3
"""Explore specific products in sample files."""

import sys
from pathlib import Path

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from openpyxl import load_workbook

from app.parser.sheet_detector import find_header_row
from app.parser.merged_cells import fill_merged_regions
from app.parser.column_mapper import map_columns
from app.parser.row_extractor import iter_product_rows
from app.parser.field_parser import parse_kv_block
from app.parser.normalizers import parse_dimensions


def explore_sample1():
    """Explore FCA-01 A in Sample1."""
    print("=" * 60)
    print("Sample1 - Looking for FCA-01 A and its WIDTH field")
    print("=" * 60)

    wb = load_workbook(project_root / 'data' / 'schedule_sample1.xlsx', data_only=False)
    ws = wb['APARTMENTS']

    header_row = find_header_row(ws)
    print(f"Header row: {header_row}")

    # Look at row 7 specifically (FCA-01 A)
    row = 7
    print(f"\nRow {row} content:")
    for col in range(1, 8):
        val = ws.cell(row=row, column=col).value
        if val:
            val_str = str(val).replace('\n', '\\n')
            print(f"  Col {col}: {val_str[:200]}...")

    # Full specs text
    specs = ws.cell(row=row, column=4).value
    print(f"\nFull SPECIFICATIONS text:")
    print(specs)

    # Parse and show parsed KV
    kv = parse_kv_block(specs)
    print(f"\nParsed KV from specs:")
    for k, v in kv.items():
        print(f"  {k}: {v}")

    # Parse dimensions
    dims = parse_dimensions(specs)
    print(f"\nParsed dimensions from specs: {dims}")


def explore_sample2():
    """Explore FTI-01 A in Sample2."""
    print("\n" + "=" * 60)
    print("Sample2 - Looking for FTI-01 A and its SIZE field")
    print("=" * 60)

    wb = load_workbook(project_root / 'data' / 'schedule_sample2.xlsx', data_only=False)
    ws = wb['Schedule']

    header_row = find_header_row(ws)
    print(f"Header row: {header_row}")

    # Find FTI-01 A row
    for row in range(header_row + 1, min(30, ws.max_row + 1)):
        doc_code = ws.cell(row=row, column=1).value
        if doc_code and 'FTI-01 A' in str(doc_code):
            print(f"\nFound FTI-01 A at row {row}")

            for col in range(1, 7):
                val = ws.cell(row=row, column=col).value
                if val:
                    val_str = str(val).replace('\n', '\\n')
                    print(f"  Col {col}: {val_str[:200]}...")

            # Full specs text
            specs = ws.cell(row=row, column=4).value
            print(f"\nFull SPECIFICATIONS text:")
            print(specs)

            # Parse and show parsed KV
            kv = parse_kv_block(specs)
            print(f"\nParsed KV from specs:")
            for k, v in kv.items():
                print(f"  {k}: {v}")

            # Parse dimensions
            dims = parse_dimensions(specs)
            print(f"\nParsed dimensions from specs: {dims}")
            break


def explore_sample3():
    """Explore F88 in Sample3."""
    print("\n" + "=" * 60)
    print("Sample3 - Looking for F88 and its Size field")
    print("=" * 60)

    wb = load_workbook(project_root / 'data' / 'schedule_sample3.xlsx', data_only=False)
    ws = wb['Schedule']

    fill_merged_regions(ws)

    header_row = find_header_row(ws)
    print(f"Header row: {header_row}")

    col_map = map_columns(ws, header_row)
    print(f"Column mapping: {col_map}")

    # Find F88 row and its detail rows
    for row in range(header_row + 1, min(ws.max_row + 1, 500)):
        doc_code = ws.cell(row=row, column=1).value
        if doc_code and str(doc_code).strip() == 'F88':
            print(f"\nFound F88 at row {row}")

            # Print this row and next 10 rows
            for r in range(row, min(row + 12, ws.max_row + 1)):
                vals = []
                for col in range(1, 10):
                    val = ws.cell(row=r, column=col).value
                    if val:
                        vals.append(f"C{col}={str(val)[:40].replace(chr(10), ' ')}")
                print(f"  Row {r}: {vals}")

            break

    # Also show what iter_product_rows extracts for F88
    print("\n\nExtracting F88 via iter_product_rows:")
    for row_data in iter_product_rows(ws, header_row, col_map):
        if row_data.get('doc_code') == 'F88':
            print(f"  Row data keys: {list(row_data.keys())}")
            print(f"  doc_code: {row_data.get('doc_code')}")
            print(f"  item_name: {row_data.get('item_name')}")
            print(f"  detail_rows:")
            for dr in row_data.get('detail_rows', []):
                print(f"    {dr}")
            break


if __name__ == '__main__':
    explore_sample1()
    explore_sample2()
    explore_sample3()
