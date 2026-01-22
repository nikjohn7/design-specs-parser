#!/usr/bin/env python3
"""Verification script for Phase 3: Row Extraction + Field Parsing.

This script tests the expected values from TASKS.md section 3.7:
- Sample1 row 7: doc_code=FCA-01 A, product_name=ICONIC, colour=SILVER SHADOW, width=3660
- Sample2 row 14: doc_code=FTI-01 A, product_name=BLINK, colour=BLANCO, width=600, height=600
- Sample3 row block: doc_code=F88, brand=Eaglestone, product_name=Rectangular plinth coffee table,
                     width=1200, length=800, height=330
- Section context appears in product_description
"""

import sys
from pathlib import Path

# Add the project root to the path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from openpyxl import load_workbook

from app.parser.workbook import load_workbook_safe, get_schedule_name
from app.parser.sheet_detector import find_header_row, is_schedule_sheet, get_schedule_sheets
from app.parser.merged_cells import fill_merged_regions
from app.parser.column_mapper import map_columns, get_column_mapping_details
from app.parser.row_extractor import iter_product_rows, extract_all_products
from app.parser.field_parser import parse_kv_block, extract_product_fields
from app.core.models import Product


def load_and_parse_sample(sample_path: Path) -> tuple[str, list[Product]]:
    """Load a sample file and parse all products."""
    print(f"\n{'='*60}")
    print(f"Loading: {sample_path.name}")
    print('='*60)

    with open(sample_path, 'rb') as f:
        file_bytes = f.read()

    wb = load_workbook_safe(file_bytes)
    schedule_name = get_schedule_name(wb, sample_path.name)
    print(f"Schedule name: {schedule_name}")

    all_products: list[Product] = []

    # Get schedule sheets
    schedule_sheets = get_schedule_sheets(wb)
    print(f"Found {len(schedule_sheets)} schedule sheet(s)")

    for sheet_name, ws, header_row in schedule_sheets:
        print(f"\n  Sheet: '{sheet_name}' (header at row {header_row})")

        # Fill merged regions
        fill_merged_regions(ws)

        # Map columns
        col_map = map_columns(ws, header_row)
        print(f"  Column mapping: {col_map}")

        # Extract products
        for row_data in iter_product_rows(ws, header_row, col_map):
            # Parse KV blocks from specs and manufacturer columns
            specs_text = row_data.get('specs', '')
            manufacturer_text = row_data.get('manufacturer', '')

            kv_specs = parse_kv_block(specs_text) if specs_text else {}
            kv_manufacturer = parse_kv_block(manufacturer_text) if manufacturer_text else {}

            # Extract product fields
            product = extract_product_fields(row_data, kv_specs, kv_manufacturer)
            all_products.append(product)

    print(f"\nTotal products extracted: {len(all_products)}")
    return schedule_name, all_products


def find_product_by_doc_code(products: list[Product], doc_code: str) -> Product | None:
    """Find a product by doc_code."""
    for p in products:
        if p.doc_code == doc_code:
            return p
    return None


def verify_sample1(products: list[Product]) -> bool:
    """Verify Sample1 expected values: FCA-01 A"""
    print("\n" + "-"*50)
    print("Verifying Sample1: FCA-01 A")
    print("-"*50)

    expected = {
        'doc_code': 'FCA-01 A',
        'product_name': 'ICONIC',
        'colour': 'SILVER SHADOW',
        'width': 3660,
    }

    product = find_product_by_doc_code(products, 'FCA-01 A')
    if not product:
        print("  FAIL: Product FCA-01 A not found")
        # List first 10 doc_codes
        print("  Available doc_codes (first 10):")
        for p in products[:10]:
            print(f"    - {p.doc_code}")
        return False

    print(f"  Found product: {product.doc_code}")
    print(f"    product_name: {product.product_name}")
    print(f"    colour: {product.colour}")
    print(f"    width: {product.width}")
    print(f"    product_description: {product.product_description}")

    passed = True
    for key, expected_val in expected.items():
        actual = getattr(product, key)
        if actual != expected_val:
            print(f"  FAIL: {key} expected '{expected_val}', got '{actual}'")
            passed = False
        else:
            print(f"  PASS: {key} = '{actual}'")

    # Check section context in product_description
    if product.product_description and 'FLOORING' in product.product_description.upper():
        print(f"  PASS: Section context 'FLOORING' found in product_description")
    else:
        print(f"  WARN: Section context not found in product_description: {product.product_description}")

    return passed


def verify_sample2(products: list[Product]) -> bool:
    """Verify Sample2 expected values: FTI-01 A"""
    print("\n" + "-"*50)
    print("Verifying Sample2: FTI-01 A")
    print("-"*50)

    expected = {
        'doc_code': 'FTI-01 A',
        'product_name': 'BLINK',
        'colour': 'BLANCO',
        'width': 600,
        'height': 600,
    }

    product = find_product_by_doc_code(products, 'FTI-01 A')
    if not product:
        print("  FAIL: Product FTI-01 A not found")
        # List first 10 doc_codes
        print("  Available doc_codes (first 10):")
        for p in products[:10]:
            print(f"    - {p.doc_code}")
        return False

    print(f"  Found product: {product.doc_code}")
    print(f"    product_name: {product.product_name}")
    print(f"    colour: {product.colour}")
    print(f"    width: {product.width}")
    print(f"    height: {product.height}")

    passed = True
    for key, expected_val in expected.items():
        actual = getattr(product, key)
        if actual != expected_val:
            print(f"  FAIL: {key} expected '{expected_val}', got '{actual}'")
            passed = False
        else:
            print(f"  PASS: {key} = '{actual}'")

    return passed


def verify_sample3(products: list[Product]) -> bool:
    """Verify Sample3 expected values: F88"""
    print("\n" + "-"*50)
    print("Verifying Sample3: F88")
    print("-"*50)

    expected = {
        'doc_code': 'F88',
        'brand': 'Eaglestone',
        'product_name': 'Rectangular plinth coffee table',
        'width': 1200,
        'length': 800,
        'height': 330,
    }

    product = find_product_by_doc_code(products, 'F88')
    if not product:
        print("  FAIL: Product F88 not found")
        # List first 10 doc_codes
        print("  Available doc_codes (first 10):")
        for p in products[:10]:
            print(f"    - {p.doc_code}")
        return False

    print(f"  Found product: {product.doc_code}")
    print(f"    product_name: {product.product_name}")
    print(f"    brand: {product.brand}")
    print(f"    width: {product.width}")
    print(f"    length: {product.length}")
    print(f"    height: {product.height}")

    passed = True
    for key, expected_val in expected.items():
        actual = getattr(product, key)
        if actual != expected_val:
            print(f"  FAIL: {key} expected '{expected_val}', got '{actual}'")
            passed = False
        else:
            print(f"  PASS: {key} = '{actual}'")

    return passed


def explore_sample_data(sample_path: Path) -> None:
    """Explore sample data to understand its structure."""
    print(f"\n{'='*60}")
    print(f"Exploring: {sample_path.name}")
    print('='*60)

    wb = load_workbook(sample_path, data_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  Sheet: '{sheet_name}' (max_row={ws.max_row}, max_col={ws.max_column})")

        header_row = find_header_row(ws)
        if header_row:
            print(f"    Header row: {header_row}")

            # Print header row content
            headers = []
            for col in range(1, min(15, (ws.max_column or 1) + 1)):
                val = ws.cell(row=header_row, column=col).value
                headers.append(str(val)[:20] if val else '')
            print(f"    Headers: {headers}")

            # Print first few data rows
            print(f"    First few rows after header:")
            for row in range(header_row + 1, min(header_row + 8, (ws.max_row or header_row) + 1)):
                vals = []
                for col in range(1, min(8, (ws.max_column or 1) + 1)):
                    val = ws.cell(row=row, column=col).value
                    if val:
                        val_str = str(val)[:30].replace('\n', ' ')
                        vals.append(val_str)
                    else:
                        vals.append('')
                print(f"      Row {row}: {vals}")


def main():
    """Run all verification tests."""
    data_dir = project_root / 'data'

    sample1_path = data_dir / 'schedule_sample1.xlsx'
    sample2_path = data_dir / 'schedule_sample2.xlsx'
    sample3_path = data_dir / 'schedule_sample3.xlsx'

    results = {}

    # First, explore the data to understand structure
    if '--explore' in sys.argv:
        explore_sample_data(sample1_path)
        explore_sample_data(sample2_path)
        explore_sample_data(sample3_path)
        return

    # Parse and verify Sample1
    try:
        _, products1 = load_and_parse_sample(sample1_path)
        results['sample1'] = verify_sample1(products1)
    except Exception as e:
        print(f"\nERROR parsing Sample1: {e}")
        import traceback
        traceback.print_exc()
        results['sample1'] = False

    # Parse and verify Sample2
    try:
        _, products2 = load_and_parse_sample(sample2_path)
        results['sample2'] = verify_sample2(products2)
    except Exception as e:
        print(f"\nERROR parsing Sample2: {e}")
        import traceback
        traceback.print_exc()
        results['sample2'] = False

    # Parse and verify Sample3
    try:
        _, products3 = load_and_parse_sample(sample3_path)
        results['sample3'] = verify_sample3(products3)
    except Exception as e:
        print(f"\nERROR parsing Sample3: {e}")
        import traceback
        traceback.print_exc()
        results['sample3'] = False

    # Summary
    print("\n" + "="*60)
    print("VERIFICATION SUMMARY")
    print("="*60)

    all_passed = True
    for sample, passed in results.items():
        status = "PASS" if passed else "FAIL"
        print(f"  {sample}: {status}")
        if not passed:
            all_passed = False

    print("\n" + ("ALL TESTS PASSED!" if all_passed else "SOME TESTS FAILED"))

    return 0 if all_passed else 1


if __name__ == '__main__':
    sys.exit(main())
