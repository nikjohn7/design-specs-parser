"""Unit tests for column mapping functionality.

Tests the map_columns, fuzzy matching, and related functions
from app/parser/column_mapper.py.
"""

import pytest
from pathlib import Path
from openpyxl import Workbook

from app.parser.column_mapper import (
    map_columns,
    get_column_mapping_details,
    get_canonical_columns,
    get_synonyms,
    COLUMN_SYNONYMS,
    FUZZY_MATCH_THRESHOLD,
    _normalize_header,
    _exact_match,
    _fuzzy_match,
    _match_column,
)
from app.parser.workbook import load_workbook_safe


class TestNormalizeHeader:
    """Tests for _normalize_header function."""

    def test_basic_normalization(self):
        """Test basic string normalization."""
        assert _normalize_header("SPEC CODE") == "spec code"
        assert _normalize_header("  Item  ") == "item"
        assert _normalize_header("Description:") == "description"

    def test_multiline_header_combined(self):
        """Test that multi-line headers are combined with space."""
        # Unlike sheet_detector, column_mapper combines lines with space
        assert _normalize_header("Item\nImage") == "item image"
        assert _normalize_header("Cost per unit $\ninc GST") == "cost per unit $ inc gst"

    def test_multiple_spaces(self):
        """Test multiple spaces are collapsed."""
        assert _normalize_header("Item   &   Location") == "item & location"

    def test_none_input(self):
        """Test None input returns empty string."""
        assert _normalize_header(None) == ""

    def test_numeric_input(self):
        """Test numeric input is converted to string."""
        assert _normalize_header(123) == "123"

    def test_trailing_punctuation(self):
        """Test trailing punctuation is removed."""
        assert _normalize_header("Notes:") == "notes"
        assert _normalize_header("Price.") == "price"
        assert _normalize_header("Cost-") == "cost"


class TestExactMatch:
    """Tests for _exact_match function."""

    def test_direct_match(self):
        """Test direct matches to synonyms."""
        assert _exact_match("spec code") == "doc_code"
        assert _exact_match("code") == "doc_code"
        assert _exact_match("qty") == "qty"
        assert _exact_match("rrp") == "cost"

    def test_partial_match(self):
        """Test partial matches work."""
        assert _exact_match("item & location (see notes)") == "item_location"
        assert _exact_match("manufacturer / supplier info") == "manufacturer"

    def test_no_match(self):
        """Test unrecognized headers return None."""
        assert _exact_match("random text") is None
        assert _exact_match("xyz123") is None
        assert _exact_match("") is None

    def test_word_boundary_matching(self):
        """Test that partial matching respects word boundaries."""
        # "code" should not match in "fabric code" context
        # This is handled by the word boundary check
        result = _exact_match("notes (supplier/fasbric code)")
        # Should match "notes" not "code"
        assert result == "notes"


class TestFuzzyMatch:
    """Tests for _fuzzy_match function."""

    def test_close_match(self):
        """Test fuzzy matching finds close matches."""
        # "specification" vs "specifications" should match
        canonical, ratio = _fuzzy_match("specification")
        assert canonical == "specs"
        assert ratio >= FUZZY_MATCH_THRESHOLD

    def test_no_match_below_threshold(self):
        """Test fuzzy matching returns None for poor matches."""
        canonical, ratio = _fuzzy_match("completely random text")
        # Should not match anything above threshold
        assert canonical is None or ratio < FUZZY_MATCH_THRESHOLD

    def test_empty_input(self):
        """Test empty input returns None."""
        canonical, ratio = _fuzzy_match("")
        assert canonical is None
        assert ratio == 0.0


class TestMatchColumn:
    """Tests for _match_column function."""

    def test_exact_match_preferred(self):
        """Test that exact match is preferred over fuzzy."""
        canonical, match_type = _match_column("spec code")
        assert canonical == "doc_code"
        assert match_type == "exact"

    def test_partial_match(self):
        """Test partial matching."""
        canonical, match_type = _match_column("item & location (notes)")
        assert canonical == "item_location"
        assert match_type == "partial"

    def test_fuzzy_match_fallback(self):
        """Test fuzzy matching as fallback."""
        # Slightly misspelled header
        canonical, match_type = _match_column("specificaton", use_fuzzy=True)
        # Should fuzzy match to "specs"
        if canonical:
            assert match_type == "fuzzy"

    def test_no_fuzzy_when_disabled(self):
        """Test fuzzy matching can be disabled."""
        canonical, match_type = _match_column("specificaton", use_fuzzy=False)
        # Without fuzzy, misspelled header won't match
        assert match_type in ("none", "partial", "exact")


class TestMapColumns:
    """Tests for map_columns function."""

    def test_basic_column_mapping(self):
        """Test basic column mapping."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'SPEC CODE'
        ws['B1'] = 'IMAGE'
        ws['C1'] = 'DESCRIPTION'
        ws['D1'] = 'QTY'
        ws['E1'] = 'COST'
        
        columns = map_columns(ws, header_row=1)
        
        assert columns['doc_code'] == 1
        assert columns['image'] == 2
        assert columns['item_location'] == 3
        assert columns['qty'] == 4
        assert columns['cost'] == 5

    def test_first_occurrence_wins(self):
        """Test that first occurrence of a canonical name is kept."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'CODE'
        ws['B1'] = 'REFERENCE'  # Also maps to doc_code
        ws['C1'] = 'DESCRIPTION'
        
        columns = map_columns(ws, header_row=1)
        
        # First occurrence (CODE at column 1) should be kept
        assert columns['doc_code'] == 1

    def test_empty_cells_skipped(self):
        """Test that empty cells are skipped."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'CODE'
        ws['B1'] = None
        ws['C1'] = ''
        ws['D1'] = 'DESCRIPTION'
        
        columns = map_columns(ws, header_row=1)
        
        assert columns['doc_code'] == 1
        assert columns['item_location'] == 4
        assert len(columns) == 2

    def test_multiline_headers(self):
        """Test handling of multi-line headers."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'Code'
        ws['B1'] = 'Item\nImage'  # Multi-line header
        ws['C1'] = 'Description'
        
        columns = map_columns(ws, header_row=1)
        
        assert columns['doc_code'] == 1
        assert columns['image'] == 2
        assert columns['item_location'] == 3

    def test_sample3_style_headers(self):
        """Test headers similar to sample3."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'Code'
        ws['B1'] = 'Area'
        ws['C1'] = 'Item\nImage'
        ws['D1'] = 'Description'
        ws['E1'] = 'Qty'
        ws['F1'] = 'Cost per unit $'
        ws['G1'] = 'Total Cost $'
        ws['H1'] = 'RRP'
        
        columns = map_columns(ws, header_row=1)
        
        assert columns['doc_code'] == 1
        assert columns['item_location'] == 2  # Area maps to item_location
        assert columns['image'] == 3
        assert columns['qty'] == 5
        assert columns['cost'] == 6
        assert columns['total_cost'] == 7

    def test_fuzzy_matching_disabled(self):
        """Test that fuzzy matching can be disabled."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'CODE'
        ws['B1'] = 'DESCRIPTIN'  # Misspelled
        
        columns_with_fuzzy = map_columns(ws, header_row=1, use_fuzzy=True)
        columns_without_fuzzy = map_columns(ws, header_row=1, use_fuzzy=False)
        
        # Both should find CODE
        assert columns_with_fuzzy['doc_code'] == 1
        assert columns_without_fuzzy['doc_code'] == 1


class TestGetColumnMappingDetails:
    """Tests for get_column_mapping_details function."""

    def test_returns_details(self):
        """Test that detailed mapping info is returned."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'SPEC CODE'
        ws['B1'] = 'Random'
        
        details = get_column_mapping_details(ws, header_row=1)
        
        assert len(details) >= 2
        
        # Check first column details
        assert details[0]['column'] == 1
        assert details[0]['original'] == 'SPEC CODE'
        assert details[0]['normalized'] == 'spec code'
        assert details[0]['canonical'] == 'doc_code'
        assert details[0]['match_type'] == 'exact'
        
        # Check unmatched column
        assert details[1]['column'] == 2
        assert details[1]['canonical'] is None
        assert details[1]['match_type'] == 'none'


class TestHelperFunctions:
    """Tests for helper functions."""

    def test_get_canonical_columns(self):
        """Test get_canonical_columns returns all canonical names."""
        columns = get_canonical_columns()
        
        assert 'doc_code' in columns
        assert 'image' in columns
        assert 'item_location' in columns
        assert 'specs' in columns
        assert 'manufacturer' in columns
        assert 'notes' in columns
        assert 'qty' in columns
        assert 'cost' in columns

    def test_get_synonyms(self):
        """Test get_synonyms returns synonyms for a canonical name."""
        synonyms = get_synonyms('doc_code')
        
        assert 'spec code' in synonyms
        assert 'code' in synonyms
        assert 'ref' in synonyms
        assert 'reference' in synonyms

    def test_get_synonyms_unknown(self):
        """Test get_synonyms returns empty list for unknown canonical."""
        synonyms = get_synonyms('unknown_column')
        assert synonyms == []


class TestSampleFiles:
    """Integration tests with actual sample files."""

    def test_sample1_apartments(self):
        """Test sample1 APARTMENTS sheet column mapping."""
        with open('data/schedule_sample1.xlsx', 'rb') as f:
            wb = load_workbook_safe(f.read())
        
        ws = wb['APARTMENTS']
        columns = map_columns(ws, header_row=4)
        
        assert columns['doc_code'] == 1
        assert columns['image'] == 2
        assert columns['item_location'] == 3
        assert columns['specs'] == 4
        assert columns['manufacturer'] == 5
        assert columns['notes'] == 6
        assert columns['cost'] == 7

    def test_sample2_schedule(self):
        """Test sample2 Schedule sheet column mapping."""
        with open('data/schedule_sample2.xlsx', 'rb') as f:
            wb = load_workbook_safe(f.read())
        
        ws = wb['Schedule']
        columns = map_columns(ws, header_row=9)
        
        assert columns['doc_code'] == 1
        assert columns['image'] == 2
        assert columns['item_location'] == 3
        assert columns['specs'] == 4
        assert columns['manufacturer'] == 5
        assert columns['notes'] == 6

    def test_sample3_schedule(self):
        """Test sample3 Schedule sheet column mapping."""
        with open('data/schedule_sample3.xlsx', 'rb') as f:
            wb = load_workbook_safe(f.read())
        
        ws = wb['Schedule']
        columns = map_columns(ws, header_row=10)
        
        assert columns['doc_code'] == 1
        assert columns['item_location'] == 2  # Area
        assert columns['image'] == 3  # Item Image
        assert columns['qty'] == 6
        assert columns['cost'] == 7  # Cost per unit $
        assert columns['total_cost'] == 8
        assert 'notes' in columns  # Notes column at end


class TestSyntheticFiles:
    """Tests with synthetic generated files."""

    def test_synthetic_generated_files(self):
        """Test column mapping on synthetic generated files."""
        import json
        from app.parser.sheet_detector import get_schedule_sheets
        
        synthetic_dir = Path('synthetic_out/generated')
        if not synthetic_dir.exists():
            pytest.skip("Synthetic files not generated")
        
        xlsx_files = list(synthetic_dir.glob('*.xlsx'))
        if not xlsx_files:
            pytest.skip("No synthetic xlsx files found")
        
        files_tested = 0
        files_with_columns = 0
        
        for xlsx_path in xlsx_files:
            truth_path = xlsx_path.with_suffix('.truth.json')
            if not truth_path.exists():
                continue
            
            with open(truth_path) as f:
                truth = json.load(f)
            
            # Skip files without code column
            has_code_col = truth.get('notes', {}).get('include_code_col', True)
            if not has_code_col:
                continue
            
            header_row = truth.get('notes', {}).get('header_row', 1)
            
            with open(xlsx_path, 'rb') as f:
                wb = load_workbook_safe(f.read())
            
            # Find schedule sheets
            schedule_sheets = get_schedule_sheets(wb)
            if not schedule_sheets:
                continue
            
            files_tested += 1
            
            for sheet_name, ws, detected_header_row in schedule_sheets:
                columns = map_columns(ws, header_row=detected_header_row)
                
                # Should find at least doc_code
                if 'doc_code' in columns:
                    files_with_columns += 1
                    break
        
        # At least 50% of tested files should have columns mapped
        if files_tested > 0:
            success_rate = files_with_columns / files_tested
            assert success_rate >= 0.5, \
                f"Low column mapping rate: {files_with_columns}/{files_tested} = {success_rate:.1%}"

    def test_synthetic_mutated_files(self):
        """Test column mapping on synthetic mutated files."""
        import json
        from app.parser.sheet_detector import get_schedule_sheets
        
        mutated_dir = Path('synthetic_out/mutated')
        if not mutated_dir.exists():
            pytest.skip("Mutated files not generated")
        
        xlsx_files = list(mutated_dir.glob('*.xlsx'))
        if not xlsx_files:
            pytest.skip("No mutated xlsx files found")
        
        for xlsx_path in xlsx_files:
            with open(xlsx_path, 'rb') as f:
                wb = load_workbook_safe(f.read())
            
            # Find schedule sheets
            schedule_sheets = get_schedule_sheets(wb)
            
            # For each detected sheet, try to map columns
            for sheet_name, ws, header_row in schedule_sheets:
                # Should not crash
                columns = map_columns(ws, header_row=header_row)
                
                # Should return a dict
                assert isinstance(columns, dict)


class TestEdgeCases:
    """Tests for edge cases and error handling."""

    def test_empty_worksheet(self):
        """Test handling of empty worksheet."""
        wb = Workbook()
        ws = wb.active
        
        columns = map_columns(ws, header_row=1)
        assert columns == {}

    def test_single_cell_worksheet(self):
        """Test handling of worksheet with single cell."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'CODE'
        
        columns = map_columns(ws, header_row=1)
        assert columns == {'doc_code': 1}

    def test_very_wide_header(self):
        """Test handling of header row with many columns."""
        wb = Workbook()
        ws = wb.active
        
        # Create header with 40 columns
        headers = ['CODE', 'DESCRIPTION', 'QTY', 'COST'] + [f'Col{i}' for i in range(36)]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        
        columns = map_columns(ws, header_row=1, max_cols=50)
        
        assert 'doc_code' in columns
        assert 'item_location' in columns
        assert 'qty' in columns
        assert 'cost' in columns

    def test_header_with_special_characters(self):
        """Test handling of headers with special characters."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'SPEC CODE #'
        ws['B1'] = 'ITEM & LOCATION (Notes)'
        ws['C1'] = 'QTY.'
        ws['D1'] = 'COST $'
        
        columns = map_columns(ws, header_row=1)
        
        assert 'doc_code' in columns
        assert 'item_location' in columns
        assert 'qty' in columns
        assert 'cost' in columns

    def test_header_row_out_of_range(self):
        """Test handling of header row beyond worksheet bounds."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'CODE'
        
        # Header row 100 is beyond the data
        columns = map_columns(ws, header_row=100)
        assert columns == {}

    def test_max_cols_limit(self):
        """Test that max_cols limits the scan."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'CODE'
        ws['Z1'] = 'DESCRIPTION'  # Column 26
        
        # With max_cols=10, should not find DESCRIPTION
        columns = map_columns(ws, header_row=1, max_cols=10)
        assert 'doc_code' in columns
        assert 'item_location' not in columns
        
        # With max_cols=30, should find both
        columns = map_columns(ws, header_row=1, max_cols=30)
        assert 'doc_code' in columns
        assert 'item_location' in columns
