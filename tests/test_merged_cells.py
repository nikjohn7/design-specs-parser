"""Unit tests for merged cell handling functionality.

Tests the fill_merged_regions function and helper functions
from app/parser/merged_cells.py.
"""

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.merge import MergedCellRange

from app.parser.merged_cells import (
    fill_merged_regions,
    get_merged_cell_value,
    is_merged_cell,
    is_merged_cell_topleft,
    get_merged_range_for_cell,
    _cell_in_range,
)
from app.parser.workbook import load_workbook_safe


class TestFillMergedRegions:
    """Tests for fill_merged_regions function."""

    def test_single_merged_range_horizontal(self):
        """Test filling a horizontally merged range (same row, multiple columns)."""
        wb = Workbook()
        ws = wb.active
        
        # Set value in A1 and merge A1:C1
        ws['A1'] = 'Merged Value'
        ws.merge_cells('A1:C1')
        
        # Before filling, B1 and C1 should be empty/None
        assert ws['B1'].value is None
        assert ws['C1'].value is None
        
        # Fill merged regions
        fill_merged_regions(ws)
        
        # After filling, all cells should have the value
        assert ws['A1'].value == 'Merged Value'
        assert ws['B1'].value == 'Merged Value'
        assert ws['C1'].value == 'Merged Value'

    def test_single_merged_range_vertical(self):
        """Test filling a vertically merged range (same column, multiple rows)."""
        wb = Workbook()
        ws = wb.active
        
        # Set value in A1 and merge A1:A3
        ws['A1'] = 'Vertical Merge'
        ws.merge_cells('A1:A3')
        
        # Fill merged regions
        fill_merged_regions(ws)
        
        # All cells should have the value
        assert ws['A1'].value == 'Vertical Merge'
        assert ws['A2'].value == 'Vertical Merge'
        assert ws['A3'].value == 'Vertical Merge'

    def test_rectangular_merged_range(self):
        """Test filling a rectangular merged range (multiple rows and columns)."""
        wb = Workbook()
        ws = wb.active
        
        # Set value in A1 and merge A1:C3
        ws['A1'] = 'Rectangle'
        ws.merge_cells('A1:C3')
        
        # Fill merged regions
        fill_merged_regions(ws)
        
        # All 9 cells should have the value
        for row in range(1, 4):
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                assert cell.value == 'Rectangle', f"Cell ({row}, {col}) should be 'Rectangle'"

    def test_multiple_merged_ranges(self):
        """Test filling multiple merged ranges in the same worksheet."""
        wb = Workbook()
        ws = wb.active
        
        # Create multiple merged ranges
        ws['A1'] = 'First'
        ws.merge_cells('A1:B1')
        
        ws['A3'] = 'Second'
        ws.merge_cells('A3:C3')
        
        ws['E1'] = 'Third'
        ws.merge_cells('E1:E3')
        
        # Fill merged regions
        fill_merged_regions(ws)
        
        # Check first range
        assert ws['A1'].value == 'First'
        assert ws['B1'].value == 'First'
        
        # Check second range
        assert ws['A3'].value == 'Second'
        assert ws['B3'].value == 'Second'
        assert ws['C3'].value == 'Second'
        
        # Check third range
        assert ws['E1'].value == 'Third'
        assert ws['E2'].value == 'Third'
        assert ws['E3'].value == 'Third'

    def test_merged_range_with_none_value(self):
        """Test filling a merged range where top-left cell is None."""
        wb = Workbook()
        ws = wb.active
        
        # Merge without setting a value
        ws.merge_cells('A1:B2')
        
        # Fill merged regions
        fill_merged_regions(ws)
        
        # All cells should be None
        assert ws['A1'].value is None
        assert ws['A2'].value is None
        assert ws['B1'].value is None
        assert ws['B2'].value is None

    def test_merged_range_with_numeric_value(self):
        """Test filling a merged range with numeric value."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 42
        ws.merge_cells('A1:B1')
        
        fill_merged_regions(ws)
        
        assert ws['A1'].value == 42
        assert ws['B1'].value == 42

    def test_merged_range_with_formula(self):
        """Test filling a merged range with formula (formula string is copied)."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = '=SUM(D1:D10)'
        ws.merge_cells('A1:B1')
        
        fill_merged_regions(ws)
        
        # Formula string should be copied
        assert ws['A1'].value == '=SUM(D1:D10)'
        assert ws['B1'].value == '=SUM(D1:D10)'

    def test_no_merged_ranges(self):
        """Test worksheet with no merged ranges."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'Value 1'
        ws['B1'] = 'Value 2'
        
        # Should not raise any errors
        fill_merged_regions(ws)
        
        # Values should be unchanged
        assert ws['A1'].value == 'Value 1'
        assert ws['B1'].value == 'Value 2'

    def test_preserves_non_merged_cells(self):
        """Test that non-merged cells are not affected."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'Merged'
        ws.merge_cells('A1:B1')
        
        ws['C1'] = 'Not Merged'
        ws['D1'] = 'Also Not Merged'
        
        fill_merged_regions(ws)
        
        # Non-merged cells should be unchanged
        assert ws['C1'].value == 'Not Merged'
        assert ws['D1'].value == 'Also Not Merged'


class TestFillMergedRegionsWithSampleFiles:
    """Integration tests with actual sample files."""

    def test_sample1_section_headers(self):
        """Test that sample1 section headers are filled correctly."""
        with open('data/schedule_sample1.xlsx', 'rb') as f:
            wb = load_workbook_safe(f.read())
        
        ws = wb.active
        
        # Before filling, check that merged cells exist
        merged_count = len(ws.merged_cells.ranges)
        assert merged_count > 0, "Sample1 should have merged cells"
        
        # Fill merged regions
        fill_merged_regions(ws)
        
        # Check FLOORING section header (A6:F6)
        # After filling, all cells in the range should have the value
        flooring_value = ws['A6'].value
        assert flooring_value == 'FLOORING'
        assert ws['B6'].value == 'FLOORING'
        assert ws['C6'].value == 'FLOORING'
        
        # Check title row (A1:F1)
        title_value = ws['A1'].value
        assert '12006' in str(title_value) or 'GEM' in str(title_value)
        assert ws['B1'].value == title_value
        assert ws['C1'].value == title_value

    def test_sample2_minimal_merges(self):
        """Test sample2 which has minimal merged cells."""
        with open('data/schedule_sample2.xlsx', 'rb') as f:
            wb = load_workbook_safe(f.read())
        
        # Check Schedule sheet
        ws = wb['Schedule']
        
        # Fill merged regions
        fill_merged_regions(ws)
        
        # Should not raise any errors
        # Check that the disclaimer row is filled
        disclaimer_value = ws['A8'].value
        if disclaimer_value:
            assert ws['B8'].value == disclaimer_value

    def test_sample3_many_merges(self):
        """Test sample3 which has many merged cells (1234)."""
        with open('data/schedule_sample3.xlsx', 'rb') as f:
            wb = load_workbook_safe(f.read())
        
        ws = wb.active
        
        # Count merged ranges before
        merged_count = len(ws.merged_cells.ranges)
        assert merged_count > 1000, f"Sample3 should have many merged cells, got {merged_count}"
        
        # Fill merged regions (should handle large number efficiently)
        fill_merged_regions(ws)
        
        # Spot check: A250:A256 should have 'F90'
        # After filling, all cells in the range should have the value
        f90_value = ws['A250'].value
        if f90_value == 'F90':
            assert ws['A251'].value == 'F90'
            assert ws['A252'].value == 'F90'


class TestGetMergedCellValue:
    """Tests for get_merged_cell_value function."""

    def test_top_left_cell(self):
        """Test getting value from top-left cell of merged range."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'Test Value'
        ws.merge_cells('A1:C3')
        
        value = get_merged_cell_value(ws, 1, 1)
        assert value == 'Test Value'

    def test_middle_cell_of_merged_range(self):
        """Test getting value from middle cell of merged range."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'Test Value'
        ws.merge_cells('A1:C3')
        
        # B2 is in the middle of the merged range
        value = get_merged_cell_value(ws, 2, 2)
        assert value == 'Test Value'

    def test_bottom_right_cell_of_merged_range(self):
        """Test getting value from bottom-right cell of merged range."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'Test Value'
        ws.merge_cells('A1:C3')
        
        # C3 is the bottom-right of the merged range
        value = get_merged_cell_value(ws, 3, 3)
        assert value == 'Test Value'

    def test_non_merged_cell(self):
        """Test getting value from non-merged cell."""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 'Merged'
        ws.merge_cells('A1:B1')
        
        ws['C1'] = 'Not Merged'
        
        value = get_merged_cell_value(ws, 1, 3)
        assert value == 'Not Merged'

    def test_empty_non_merged_cell(self):
        """Test getting value from empty non-merged cell."""
        wb = Workbook()
        ws = wb.active
        
        value = get_merged_cell_value(ws, 5, 5)
        assert value is None


class TestIsMergedCell:
    """Tests for is_merged_cell function."""

    def test_cell_in_merged_range(self):
        """Test that cells in merged range are identified."""
        wb = Workbook()
        ws = wb.active
        
        ws.merge_cells('A1:C3')
        
        # All cells in the range should be identified as merged
        assert is_merged_cell(ws, 1, 1) is True
        assert is_merged_cell(ws, 1, 2) is True
        assert is_merged_cell(ws, 2, 2) is True
        assert is_merged_cell(ws, 3, 3) is True

    def test_cell_not_in_merged_range(self):
        """Test that cells outside merged range are not identified."""
        wb = Workbook()
        ws = wb.active
        
        ws.merge_cells('A1:B2')
        
        # Cells outside the range should not be identified as merged
        assert is_merged_cell(ws, 1, 3) is False
        assert is_merged_cell(ws, 3, 1) is False
        assert is_merged_cell(ws, 5, 5) is False

    def test_no_merged_cells(self):
        """Test worksheet with no merged cells."""
        wb = Workbook()
        ws = wb.active
        
        assert is_merged_cell(ws, 1, 1) is False


class TestIsMergedCellTopleft:
    """Tests for is_merged_cell_topleft function."""

    def test_top_left_cell(self):
        """Test that top-left cell is correctly identified."""
        wb = Workbook()
        ws = wb.active
        
        ws.merge_cells('B2:D4')
        
        assert is_merged_cell_topleft(ws, 2, 2) is True

    def test_non_top_left_merged_cell(self):
        """Test that non-top-left merged cells are not identified as top-left."""
        wb = Workbook()
        ws = wb.active
        
        ws.merge_cells('A1:C3')
        
        # Only A1 should be top-left
        assert is_merged_cell_topleft(ws, 1, 1) is True
        assert is_merged_cell_topleft(ws, 1, 2) is False
        assert is_merged_cell_topleft(ws, 2, 1) is False
        assert is_merged_cell_topleft(ws, 2, 2) is False

    def test_non_merged_cell(self):
        """Test that non-merged cells are not identified as top-left."""
        wb = Workbook()
        ws = wb.active
        
        ws.merge_cells('A1:B1')
        
        assert is_merged_cell_topleft(ws, 1, 3) is False
        assert is_merged_cell_topleft(ws, 2, 1) is False


class TestGetMergedRangeForCell:
    """Tests for get_merged_range_for_cell function."""

    def test_cell_in_merged_range(self):
        """Test getting merged range for cell in merged region."""
        wb = Workbook()
        ws = wb.active
        
        ws.merge_cells('A1:C3')
        
        merged_range = get_merged_range_for_cell(ws, 2, 2)
        assert merged_range is not None
        
        min_col, min_row, max_col, max_row = merged_range.bounds
        assert min_row == 1
        assert min_col == 1
        assert max_row == 3
        assert max_col == 3

    def test_cell_not_in_merged_range(self):
        """Test getting merged range for cell not in merged region."""
        wb = Workbook()
        ws = wb.active
        
        ws.merge_cells('A1:B2')
        
        merged_range = get_merged_range_for_cell(ws, 5, 5)
        assert merged_range is None


class TestCellInRange:
    """Tests for _cell_in_range helper function."""

    def test_cell_inside_range(self):
        """Test cell inside range bounds."""
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('A1:C3')
        
        merged_range = list(ws.merged_cells.ranges)[0]
        
        assert _cell_in_range(1, 1, merged_range) is True
        assert _cell_in_range(2, 2, merged_range) is True
        assert _cell_in_range(3, 3, merged_range) is True

    def test_cell_outside_range(self):
        """Test cell outside range bounds."""
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('A1:C3')
        
        merged_range = list(ws.merged_cells.ranges)[0]
        
        assert _cell_in_range(0, 1, merged_range) is False
        assert _cell_in_range(4, 1, merged_range) is False
        assert _cell_in_range(1, 4, merged_range) is False


class TestSyntheticFiles:
    """Tests with synthetic generated files."""

    def test_synthetic_generated_files(self):
        """Test fill_merged_regions on synthetic generated files."""
        import os
        from pathlib import Path
        
        synthetic_dir = Path('synthetic_out/generated')
        if not synthetic_dir.exists():
            pytest.skip("Synthetic files not generated")
        
        xlsx_files = list(synthetic_dir.glob('*.xlsx'))
        if not xlsx_files:
            pytest.skip("No synthetic xlsx files found")
        
        for xlsx_path in xlsx_files[:5]:  # Test first 5 files
            with open(xlsx_path, 'rb') as f:
                wb = load_workbook_safe(f.read())
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Should not raise any errors
                fill_merged_regions(ws)
                
                # Verify merged cells are filled
                for merged_range in ws.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    top_left_value = ws.cell(row=min_row, column=min_col).value
                    
                    # All cells in range should have the same value
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            cell_value = ws.cell(row=row, column=col).value
                            assert cell_value == top_left_value, \
                                f"Cell ({row}, {col}) in {xlsx_path.name} should have value {top_left_value}"

    def test_synthetic_mutated_files(self):
        """Test fill_merged_regions on synthetic mutated files."""
        import os
        from pathlib import Path
        
        mutated_dir = Path('synthetic_out/mutated')
        if not mutated_dir.exists():
            pytest.skip("Mutated files not generated")
        
        xlsx_files = list(mutated_dir.glob('*.xlsx'))
        if not xlsx_files:
            pytest.skip("No mutated xlsx files found")
        
        for xlsx_path in xlsx_files:
            with open(xlsx_path, 'rb') as f:
                wb = load_workbook_safe(f.read())
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Should not raise any errors
                fill_merged_regions(ws)
