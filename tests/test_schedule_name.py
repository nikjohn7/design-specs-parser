"""Unit tests for schedule name extraction functionality.

Tests the get_schedule_name function and its helper functions
from app/parser/workbook.py.
"""

import pytest
from openpyxl import Workbook

from app.parser.workbook import (
    get_schedule_name,
    load_workbook_safe,
    _is_metadata_label,
    _is_likely_title,
    _filename_to_schedule_name,
    _get_cell_string_value,
    _resolve_cover_sheet_formula,
    _find_schedule_name_in_cover_sheet,
)


class TestIsMetadataLabel:
    """Tests for _is_metadata_label helper function."""

    def test_known_metadata_labels(self):
        """Test that known metadata labels are correctly identified."""
        metadata_labels = [
            'Job No.',
            'job no.',
            'JOB NO.',
            'Job Name.',
            'Revision Date',
            'Client Name:',
            'Project Address:',
            'Version:',
            'Issue Date:',
            'Project:',
            'Notes',
            'Legend',
            'revision',
            'date',
        ]
        for label in metadata_labels:
            assert _is_metadata_label(label) is True, f"Expected '{label}' to be metadata"

    def test_non_metadata_labels(self):
        """Test that schedule titles are not identified as metadata."""
        non_metadata = [
            'Interior Schedule',
            '12006: GEM, WATERLINE PLACE',
            'SCHEDULE 003- INTERNAL FINISHES',
            'Lighting Schedule (FF&E Tracker)',
            'Apartment Schedule',
        ]
        for text in non_metadata:
            assert _is_metadata_label(text) is False, f"Expected '{text}' to NOT be metadata"

    def test_empty_and_none(self):
        """Test edge cases with empty strings."""
        assert _is_metadata_label('') is True
        assert _is_metadata_label('   ') is True

    def test_short_labels_with_colon(self):
        """Test that short labels ending with colon are metadata."""
        assert _is_metadata_label('Short:') is True
        assert _is_metadata_label('Rev:') is True

    def test_single_char(self):
        """Test single character is not metadata."""
        assert _is_metadata_label('A') is False


class TestIsLikelyTitle:
    """Tests for _is_likely_title helper function."""

    def test_schedule_titles(self):
        """Test that schedule titles are correctly identified."""
        titles = [
            '12006: GEM, WATERLINE PLACE, WILLIAMSTOWN',
            'SCHEDULE 003- INTERNAL FINISHES',
            'Interior Schedule',
            'Lighting Schedule (FF&E Tracker)',
            'PROJECT: Synthetic Interior Schedule',
            'Apartment Schedule',
            'FF&E Schedule for Project X',
        ]
        for title in titles:
            assert _is_likely_title(title) is True, f"Expected '{title}' to be a title"

    def test_non_titles(self):
        """Test that metadata labels are not identified as titles."""
        non_titles = [
            'Job No.',
            'Revision Date',
            'Client Name:',
            'Notes',
            '',
            'AB',  # Too short
        ]
        for text in non_titles:
            assert _is_likely_title(text) is False, f"Expected '{text}' to NOT be a title"

    def test_formulas_not_titles(self):
        """Test that formulas are not identified as titles."""
        formulas = [
            "='Cover Sheet'!A6",
            "=[1]Cover Sheet!A6",
            "=SUM(A1:A10)",
        ]
        for formula in formulas:
            assert _is_likely_title(formula) is False, f"Expected formula '{formula}' to NOT be a title"

    def test_error_values_not_titles(self):
        """Test that Excel error values are not identified as titles."""
        errors = [
            '#REF!',
            '#N/A!',
            '#VALUE!',
        ]
        for error in errors:
            assert _is_likely_title(error) is False, f"Expected error '{error}' to NOT be a title"


class TestFilenameToScheduleName:
    """Tests for _filename_to_schedule_name helper function."""

    def test_xlsx_extension_removal(self):
        """Test that .xlsx extension is removed."""
        assert _filename_to_schedule_name('schedule_sample1.xlsx') == 'schedule sample1'
        assert _filename_to_schedule_name('my_project.xlsx') == 'my project'

    def test_xls_extension_removal(self):
        """Test that .xls extension is removed."""
        assert _filename_to_schedule_name('test.xls') == 'test'

    def test_uppercase_extension(self):
        """Test that uppercase extensions are handled."""
        assert _filename_to_schedule_name('FILE.XLSX') == 'FILE'
        assert _filename_to_schedule_name('FILE.XLS') == 'FILE'

    def test_no_extension(self):
        """Test filenames without extension."""
        assert _filename_to_schedule_name('no_extension') == 'no extension'

    def test_empty_filename(self):
        """Test empty filename returns default."""
        assert _filename_to_schedule_name('') == 'Unknown Schedule'
        assert _filename_to_schedule_name('   ') == 'Unknown Schedule'

    def test_only_extension(self):
        """Test filename that is only extension."""
        assert _filename_to_schedule_name('.xlsx') == 'Unknown Schedule'

    def test_underscores_replaced(self):
        """Test that underscores are replaced with spaces."""
        assert _filename_to_schedule_name('multiple_underscores_here.xlsx') == 'multiple underscores here'


class TestGetScheduleNameWithWorkbook:
    """Tests for get_schedule_name with actual workbook objects."""

    def test_title_in_row_1(self):
        """Test extraction when title is in row 1."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '12006: GEM, WATERLINE PLACE, WILLIAMSTOWN'
        
        result = get_schedule_name(wb, 'test.xlsx')
        assert result == '12006: GEM, WATERLINE PLACE, WILLIAMSTOWN'

    def test_title_with_schedule_keyword(self):
        """Test extraction when title contains 'schedule' keyword."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Interior Schedule'
        
        result = get_schedule_name(wb, 'test.xlsx')
        assert result == 'Interior Schedule'

    def test_title_with_project_keyword(self):
        """Test extraction when title contains 'project' keyword."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'PROJECT: My Interior Design'
        
        result = get_schedule_name(wb, 'test.xlsx')
        assert result == 'PROJECT: My Interior Design'

    def test_fallback_to_filename(self):
        """Test fallback to filename when no title found."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Job No.'
        ws['A2'] = 'Revision Date'
        ws['A3'] = 'Notes'
        
        result = get_schedule_name(wb, 'my_schedule.xlsx')
        assert result == 'my schedule'

    def test_empty_workbook(self):
        """Test with workbook that has no sheets."""
        wb = Workbook()
        # Remove the default sheet
        for sheet in wb.sheetnames:
            del wb[sheet]
        
        result = get_schedule_name(wb, 'empty.xlsx')
        assert result == 'empty'

    def test_workbook_with_empty_cells(self):
        """Test with workbook that has only empty cells."""
        wb = Workbook()
        ws = wb.active
        # Leave all cells empty
        
        result = get_schedule_name(wb, 'empty_cells.xlsx')
        assert result == 'empty cells'

    def test_title_in_column_b(self):
        """Test extraction when title is in column B with label in A."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'SCHEDULE NAME'
        ws['B1'] = 'Interior Schedule'
        
        result = get_schedule_name(wb, 'test.xlsx')
        assert result == 'Interior Schedule'

    def test_cover_sheet_with_schedule_name_row(self):
        """Test extraction from Cover Sheet with SCHEDULE NAME label."""
        wb = Workbook()
        
        # Create Cover Sheet
        cover = wb.active
        cover.title = 'Cover Sheet'
        cover['A1'] = 'PROJECT: Test Project'
        cover['A3'] = 'SCHEDULE NAME'
        cover['B3'] = 'Interior Schedule'
        
        result = get_schedule_name(wb, 'test.xlsx')
        # Should find the PROJECT title first
        assert 'PROJECT' in result or 'Interior Schedule' in result

    def test_formula_reference_to_cover_sheet(self):
        """Test handling of formula reference to Cover Sheet."""
        wb = Workbook()
        
        # Create Cover Sheet with title
        cover = wb.active
        cover.title = 'Cover Sheet'
        cover['A6'] = 'SCHEDULE 003- INTERNAL FINISHES'
        
        # Create Schedule sheet with formula reference
        schedule = wb.create_sheet('Schedule')
        schedule['A7'] = "='[1]Cover Sheet'!A6"
        
        # Make Schedule the active sheet
        wb.active = schedule
        
        result = get_schedule_name(wb, 'test.xlsx')
        # Should resolve the formula and get the title from Cover Sheet
        assert result == 'SCHEDULE 003- INTERNAL FINISHES'


class TestGetScheduleNameWithSampleFiles:
    """Integration tests with actual sample files."""

    def test_sample1(self):
        """Test schedule name extraction from sample1."""
        with open('data/schedule_sample1.xlsx', 'rb') as f:
            wb = load_workbook_safe(f.read())
        
        result = get_schedule_name(wb, 'schedule_sample1.xlsx')
        assert result == '12006: GEM, WATERLINE PLACE, WILLIAMSTOWN'

    def test_sample2(self):
        """Test schedule name extraction from sample2 (formula reference)."""
        with open('data/schedule_sample2.xlsx', 'rb') as f:
            wb = load_workbook_safe(f.read())
        
        result = get_schedule_name(wb, 'schedule_sample2.xlsx')
        assert result == 'SCHEDULE 003- INTERNAL FINISHES'

    def test_sample3(self):
        """Test schedule name extraction from sample3 (fallback to filename)."""
        with open('data/schedule_sample3.xlsx', 'rb') as f:
            wb = load_workbook_safe(f.read())
        
        result = get_schedule_name(wb, 'schedule_sample3.xlsx')
        assert result == 'schedule sample3'


class TestEdgeCases:
    """Edge case tests for schedule name extraction."""

    def test_very_long_title(self):
        """Test handling of very long titles."""
        wb = Workbook()
        ws = wb.active
        long_title = 'A' * 500 + ' Schedule'
        ws['A1'] = long_title
        
        result = get_schedule_name(wb, 'test.xlsx')
        assert result == long_title

    def test_unicode_title(self):
        """Test handling of unicode characters in title."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Café Interior Schedule 日本語'
        
        result = get_schedule_name(wb, 'test.xlsx')
        assert 'Café' in result

    def test_numeric_title(self):
        """Test handling of numeric values in cells."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 12345
        ws['A2'] = 'Interior Schedule'
        
        result = get_schedule_name(wb, 'test.xlsx')
        assert result == 'Interior Schedule'

    def test_whitespace_only_cells(self):
        """Test handling of cells with only whitespace."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '   '
        ws['A2'] = '\t\n'
        ws['A3'] = 'Interior Schedule'
        
        result = get_schedule_name(wb, 'test.xlsx')
        assert result == 'Interior Schedule'

    def test_multiple_sheets_first_has_title(self):
        """Test with multiple sheets where first has title."""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Main'
        ws1['A1'] = 'Main Schedule'
        
        ws2 = wb.create_sheet('Other')
        ws2['A1'] = 'Other Schedule'
        
        result = get_schedule_name(wb, 'test.xlsx')
        assert result == 'Main Schedule'
