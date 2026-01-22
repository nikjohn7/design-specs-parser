"""Programa Schedule Test Data Generator

Generates and/or mutates Excel (.xlsx) product schedules to stress-test parsers
for the Programa MLE take-home challenge.

Key features
------------
- Two modes:
  1) generate: create brand-new synthetic schedules (with ground-truth JSON)
  2) mutate:   take existing .xlsx samples/templates and apply realistic
               mutations (header shifts, merged cells, hidden columns, etc.)

- Multiple layout families (sample-like finishes schedule, normalized table,
  FF&E tracker-style).

- Mutation pipeline with reproducible RNG seed.

- Optional embedded images (requires Pillow; degrades gracefully).

- Emits a manifest.jsonl in the output folder with metadata for each file.

This is intentionally "no-nonsense": it focuses on producing *varied* and
*realistic* spreadsheets that reflect common interior design schedule patterns.

Usage examples
--------------
Generate 50 new schedules (with images) and 5 mutants per sample workbook:

  python generate_programa_test_schedules.py \
    --mode both \
    --samples_dir ./data \
    --num_generated 50 \
    --mutants_per_sample 5 \
    --mutations 2-6 \
    --with_images \
    --output_dir ./synthetic_out \
    --seed 123

Mutate only:
  python generate_programa_test_schedules.py --mode mutate --samples_dir ./data --output_dir ./out

Generate only (no images):
  python generate_programa_test_schedules.py --mode generate --num_generated 100 --output_dir ./out
"""

from __future__ import annotations

import argparse
import dataclasses
import datetime as _dt
import hashlib
import json
import logging
import math
import os
import random
import re
import shutil
import string
import sys
import textwrap
import uuid
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple

# Configure logging for mutation tracking
logging.basicConfig(
    level=logging.WARNING,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# NOTE: we avoid importing MergedCell globally; some openpyxl versions move it.


# -----------------------------
# Optional Pillow for images
# -----------------------------

try:
    from PIL import Image as PILImage
    from PIL import ImageDraw

    _PIL_AVAILABLE = True
except Exception:
    _PIL_AVAILABLE = False

try:
    from openpyxl.drawing.image import Image as XLImage

    _OPENPYXL_IMAGE_AVAILABLE = True
except Exception:
    _OPENPYXL_IMAGE_AVAILABLE = False


# -----------------------------
# Data model
# -----------------------------


@dataclass
class ProductTruth:
    doc_code: str
    product_name: Optional[str] = None
    brand: Optional[str] = None
    colour: Optional[str] = None
    finish: Optional[str] = None
    material: Optional[str] = None
    width: Optional[int] = None
    length: Optional[int] = None
    height: Optional[int] = None
    qty: Optional[int] = None
    rrp: Optional[Decimal] = None
    feature_image: Optional[str] = None
    product_description: Optional[str] = None
    product_details: Optional[str] = None

    # Helpful extra metadata (not part of the parser output schema)
    _category: Optional[str] = None
    _room: Optional[str] = None


@dataclass
class ScheduleTruth:
    schedule_name: str
    products: List[ProductTruth]
    layout_family: str
    seed: int
    mutations: List[str]
    notes: Dict[str, Any]


# -----------------------------
# Header synonym pools
# -----------------------------


TARGET_HEADERS: Dict[str, List[str]] = {
    "doc_code": [
        "SPEC CODE",
        "Item Code",
        "Code",
        "Ref No",
        "Reference",
        "ID",
        "SKU",
        "Drawing Code",
        "Doc Code",
    ],
    "product_name": [
        "PRODUCT NAME",
        "Item Name",
        "Product",
        "Name",
        "Item",
        "DESCRIPTION",
    ],
    "brand": [
        "MANUFACTURER / SUPPLIER",
        "Manufacturer",
        "Supplier",
        "Brand",
        "Vendor",
        "Make",
    ],
    "colour": [
        "COLOUR",
        "Color",
        "Finish Color",
        "Colour Code",
    ],
    "finish": [
        "FINISH",
        "Surface",
        "Surface Finish",
        "Texture",
    ],
    "material": [
        "MATERIAL",
        "Materials",
        "Composition",
        "Main Material",
    ],
    "width": [
        "WIDTH (mm)",
        "W (mm)",
        "WIDTH",
        "W",
    ],
    "length": [
        "LENGTH (mm)",
        "L (mm)",
        "LENGTH",
        "DEPTH",
        "D",
        "L",
    ],
    "height": [
        "HEIGHT (mm)",
        "H (mm)",
        "HEIGHT",
        "H",
    ],
    "qty": [
        "QTY",
        "Quantity",
        "Units",
        "No.",
        "QTY (ea)",
    ],
    "rrp": [
        "RRP",
        "Price",
        "Unit Price",
        "INDICATIVE COST",
        "Cost",
        "$",
    ],
    "feature_image": [
        "IMAGE",
        "Photo",
        "Picture",
        "INDICATIVE IMAGE",
        "Indicative Image",
        "Image Link",
    ],
    "product_description": [
        "ITEM & LOCATION",
        "Product Description",
        "Description",
        "Room",
        "Location",
    ],
    "product_details": [
        "SPECIFICATIONS",
        "Specs",
        "Details",
        "NOTES",
        "COMMENTS",
        "NOTES/COMMENTS",
    ],
}


HEADER_LOOKUP: Dict[str, str] = {}
for k, names in TARGET_HEADERS.items():
    for n in names:
        HEADER_LOOKUP[re.sub(r"\s+", " ", n.strip().upper())] = k


def normalize_header_text(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().upper())


# -----------------------------
# Random data pools
# -----------------------------


ROOMS = [
    "Living Room",
    "Kitchen",
    "Bathroom",
    "Powder Room",
    "Bedroom 1",
    "Bedroom 2",
    "Laundry",
    "Entry",
    "Hallway",
    "Balcony",
    "Study",
    "Dining",
]

CATEGORIES = [
    "FLOORING",
    "WALL FINISHES",
    "PAINT",
    "LIGHTING",
    "JOINERY",
    "HARDWARE",
    "SANITARY",
    "TILES",
    "STONE",
    "GLASS",
    "FURNITURE",
    "APPLIANCES",
]

COLOURS = [
    "White",
    "Black",
    "Charcoal",
    "Warm Grey",
    "Cool Grey",
    "Natural Oak",
    "Walnut",
    "Brass",
    "Chrome",
    "Matte Nickel",
    "Off‑White",
    "Bone",
    "Écru",
    "Terracotta",
    "Sage",
    "Navy",
    "Forest Green",
]

FINISHES = [
    "Matte",
    "Matt",
    "Polished",
    "Honed",
    "Brushed",
    "Brushed Brass",
    "Powdercoat",
    "Powder‑coat",
    "Satin",
    "Gloss",
    "Textured",
    "Ribbed",
    "Natural",
]

MATERIALS = [
    "Porcelain",
    "Ceramic",
    "Timber",
    "Engineered Timber",
    "Hardwood",
    "Stone",
    "Marble",
    "Quartz",
    "Stainless Steel",
    "Mild Steel",
    "Aluminium",
    "Glass",
    "Brass",
    "Acrylic",
    "Laminate",
    "Polyurethane",
    "Wool",
    "Linen",
]

BRANDS = [
    "Timber Plus",
    "Ceramica",
    "LuxLighting Co.",
    "Studio Hardware",
    "StoneWorks",
    "GlassLab",
    "KitchenCraft",
    "BathHaus",
    "Forma",
    "Atlas",
    "Tekton",
    "Café Supply",
    "Nørd Interiors",
]

SUPPLIER_DOMAINS = [
    "example.com",
    "supplier.test",
    "studio.local",
    "vendor.invalid",
]


ADJECTIVES = [
    "Minimalist",
    "Classic",
    "Contemporary",
    "Slimline",
    "Ribbed",
    "Handmade",
    "Heritage",
    "Architectural",
    "Textured",
    "Soft",
]

PRODUCT_TYPES_BY_CATEGORY: Dict[str, List[str]] = {
    "FLOORING": ["Oak Flooring", "Engineered Timber Plank", "Carpet Tile", "Vinyl Plank"],
    "TILES": ["Porcelain Tile", "Subway Tile", "Terrazzo Tile", "Mosaic Sheet"],
    "WALL FINISHES": ["Wall Panel", "Acoustic Panel", "Feature Wallpaper"],
    "PAINT": ["Interior Paint", "Ceiling Paint", "Trim Enamel"],
    "LIGHTING": ["Pendant Light", "Wall Sconce", "Downlight", "Track Light"],
    "JOINERY": ["Cabinet Door", "Laminate Benchtop", "Stone Benchtop", "Vanity Unit"],
    "HARDWARE": ["Door Handle", "Hinge", "Pull Handle", "Floor Waste"],
    "SANITARY": ["Basin", "Tapware", "Shower Mixer", "Toilet Suite"],
    "STONE": ["Marble Slab", "Quartz Slab", "Granite Tile"],
    "GLASS": ["Shower Screen Glass", "Balustrade Glass", "Mirror"],
    "FURNITURE": ["Dining Chair", "Sofa", "Coffee Table", "Side Table"],
    "APPLIANCES": ["Oven", "Cooktop", "Rangehood", "Dishwasher"],
}


def _rng_choice(rng: random.Random, items: Sequence[Any]) -> Any:
    if not items:
        raise ValueError("Empty choice pool")
    return items[rng.randrange(len(items))]


def _maybe(rng: random.Random, p: float) -> bool:
    return rng.random() < p


def _rand_float(rng: random.Random, a: float, b: float) -> float:
    return a + (b - a) * rng.random()


def _rand_int(rng: random.Random, a: int, b: int) -> int:
    return rng.randint(a, b)


def _random_email(rng: random.Random, brand: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", ".", brand.lower()).strip(".")
    user = _rng_choice(rng, ["sales", "hello", "info", "specs", "orders"]) + str(_rand_int(rng, 1, 99))
    domain = _rng_choice(rng, SUPPLIER_DOMAINS)
    return f"{user}@{slug}.{domain}".replace("..", ".")


def _random_phone(rng: random.Random) -> str:
    # Australian-ish formatting, but not guaranteed valid.
    if _maybe(rng, 0.5):
        return f"+61 4{_rand_int(rng,0,9)}{_rand_int(rng,0,9)} {_rand_int(rng,100,999)} {_rand_int(rng,100,999)}"
    return f"(02) {_rand_int(rng,1000,9999)} {_rand_int(rng,1000,9999)}"


def _random_address(rng: random.Random) -> str:
    streets = ["George St", "Clarence St", "Collins St", "King St", "Bourke St", "Oxford St"]
    cities = ["Sydney", "Melbourne", "Brisbane", "Perth", "Adelaide"]
    return f"{_rand_int(rng, 1, 300)} {_rng_choice(rng, streets)}, {_rng_choice(rng, cities)}"


def _random_url(rng: random.Random, brand: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", brand.lower()).strip("-")
    return f"https://{slug}.{_rng_choice(rng, SUPPLIER_DOMAINS)}"


# -----------------------------
# Dimension & price formatting
# -----------------------------


def mm_to_repr(rng: random.Random, mm: Optional[int], kind: str = "generic") -> Optional[str]:
    """Represent a dimension in one of many real-world formats.

    mm: canonical dimension in millimetres.
    kind: influences likely units (e.g., flooring often uses metres).
    """

    if mm is None:
        return None
    fmt = rng.random()

    # flooring/large runs: metres more likely
    metres_bias = 0.6 if kind in {"FLOORING", "WALL FINISHES"} else 0.25
    if fmt < metres_bias:
        m = mm / 1000.0
        # match sample-like "3.66 METRES"
        return f"{m:.2f} METRES".replace(".00", ".0")
    if fmt < metres_bias + 0.20:
        cm = mm / 10.0
        return f"{cm:.0f} cm"
    if fmt < metres_bias + 0.35:
        # imperial-ish
        inches = mm / 25.4
        return f"{inches:.1f}\""  # 10.5"
    if fmt < metres_bias + 0.60:
        return f"{mm}mm"
    return str(mm)


def dims_block(
    rng: random.Random,
    w: Optional[int],
    l: Optional[int],
    h: Optional[int],
    category: str,
) -> str:
    """Create dimension lines in varied formats."""

    parts: List[str] = []
    style = rng.random()
    if style < 0.35:
        # Separate lines
        if w is not None:
            parts.append(f"WIDTH: {mm_to_repr(rng, w, category)}")
        if l is not None:
            parts.append(f"LENGTH: {mm_to_repr(rng, l, category)}")
        if h is not None:
            parts.append(f"HEIGHT: {mm_to_repr(rng, h, category)}")
    elif style < 0.70:
        # Compact W x L x H
        # Use canonical mm but vary units label.
        unit = _rng_choice(rng, ["mm", "MM", ""])
        if w is not None and l is not None and h is not None:
            parts.append(f"SIZE: {w} x {l} x {h}{unit}")
        elif w is not None and l is not None:
            parts.append(f"SIZE: {w} x {l}{unit}")
        elif w is not None:
            parts.append(f"SIZE: {w}{unit}")
    else:
        # Product-specific shorthand
        keys = _rng_choice(rng, ["W", "D", "H", "L"])
        # Keep this simple: W/D/H
        if w is not None:
            parts.append(f"W: {mm_to_repr(rng, w, category)}")
        if l is not None:
            parts.append(f"D: {mm_to_repr(rng, l, category)}")
        if h is not None:
            parts.append(f"H: {mm_to_repr(rng, h, category)}")
    return "\n".join(parts)


def price_to_repr(rng: random.Random, price: Optional[Decimal], category: str) -> Optional[str]:
    if price is None:
        return _rng_choice(rng, [None, "POA", "TBC", "N/A"])  # type: ignore[return-value]
    p = float(price)
    style = rng.random()
    currency = _rng_choice(rng, ["$", "AUD ", "USD "])
    if style < 0.40:
        return f"{currency}{p:.2f}"
    if style < 0.65:
        suffix = _rng_choice(rng, [" +GST", " + GST", " inc GST", " ex GST"])
        return f"{currency}{p:.2f}{suffix}"
    if style < 0.85:
        # Per sqm / supply & install
        per = "PER SQM" if category in {"FLOORING", "TILES", "STONE"} else "PER EA"
        extra = _rng_choice(rng, ["SUPPLY ONLY", "SUPPLY AND INSTALL", "SUPPLY + INSTALL"])
        return f"{currency}{p:.2f} +GST {per} {extra}"
    # Range
    lo = max(0.0, p * _rand_float(rng, 0.85, 0.95))
    hi = p * _rand_float(rng, 1.05, 1.25)
    return f"{currency}{lo:.0f}–{hi:.0f}"


# -----------------------------
# Spec / supplier block grammar
# -----------------------------


def spec_block(rng: random.Random, prod: ProductTruth) -> str:
    """Create multi-line, semi-structured SPECIFICATIONS content."""

    # Key variants (stress parsers)
    key_variants = {
        "PRODUCT": ["PRODUCT", "Product", "ITEM"],
        "CODE": ["CODE", "Code", "REF"],
        "COLOUR": ["COLOUR", "COLOR", "Colour"],
        "FINISH": ["FINISH", "Finish", "SURFACE"],
        "COMPOSITION": ["COMPOSITION", "MATERIAL", "Composition"],
    }
    delim = _rng_choice(rng, [": ", " - ", " = ", ":"])  # include colon-without-space

    lines: List[str] = []
    if prod.product_name and _maybe(rng, 0.9):
        lines.append(f"{_rng_choice(rng, key_variants['PRODUCT'])}{delim}{prod.product_name}")
    if _maybe(rng, 0.8):
        lines.append(f"{_rng_choice(rng, key_variants['CODE'])}{delim}{prod.doc_code}")
    if prod.colour and _maybe(rng, 0.85):
        lines.append(f"{_rng_choice(rng, key_variants['COLOUR'])}{delim}{prod.colour}")
    if prod.finish and _maybe(rng, 0.75):
        lines.append(f"{_rng_choice(rng, key_variants['FINISH'])}{delim}{prod.finish}")
    if prod.material and _maybe(rng, 0.75):
        lines.append(f"{_rng_choice(rng, key_variants['COMPOSITION'])}{delim}{prod.material}")

    # Dimensions are common but not guaranteed
    if _maybe(rng, 0.8):
        lines.append(dims_block(rng, prod.width, prod.length, prod.height, prod._category or "generic"))

    # Extra details
    extras = [
        "Refer to drawings for coordination.",
        "Install per manufacturer specification.",
        "Allow for cutting waste.",
        "Finish sample to be approved prior to order.",
        "All dimensions to be verified on site.",
        "Lead time subject to confirmation.",
    ]
    if _maybe(rng, 0.6):
        lines.append(_rng_choice(rng, extras))

    # Shuffle order sometimes
    if _maybe(rng, 0.5):
        rng.shuffle(lines)

    # Occasionally turn into a single paragraph (remove newlines)
    if _maybe(rng, 0.10):
        return " | ".join([ln.replace("\n", " ") for ln in lines if ln])

    # Ensure we don't return empty
    return "\n".join([ln for ln in lines if ln]).strip() or (prod.product_details or "")


def supplier_block(rng: random.Random, prod: ProductTruth) -> str:
    brand = prod.brand or _rng_choice(rng, BRANDS)
    contact_name = _rng_choice(rng, ["Alex", "Sam", "Taylor", "Jordan", "Casey", "Morgan"]) + " " + _rng_choice(
        rng, ["Lee", "Ng", "Patel", "Kim", "Smith", "Garcia"]
    )
    lines: List[str] = []
    role = _rng_choice(rng, ["MANUFACTURER", "SUPPLIER", "VENDOR", "MAKE"])
    delim = _rng_choice(rng, [": ", " - ", " = "])
    lines.append(f"{role}{delim}{brand}")
    if _maybe(rng, 0.7):
        lines.append(f"CONTACT{delim}{contact_name}")
    if _maybe(rng, 0.8):
        lines.append(f"WEB{delim}{_random_url(rng, brand)}")
    if _maybe(rng, 0.8):
        lines.append(f"EMAIL{delim}{_random_email(rng, brand)}")
    if _maybe(rng, 0.7):
        lines.append(f"PHONE{delim}{_random_phone(rng)}")
    if _maybe(rng, 0.5):
        lines.append(f"ADDRESS{delim}{_random_address(rng)}")
    if _maybe(rng, 0.4):
        lines.append(f"ABN{delim}{_rand_int(rng, 10, 99)} {_rand_int(rng, 100, 999)} {_rand_int(rng, 100, 999)} {_rand_int(rng, 100, 999)}")
    if _maybe(rng, 0.3):
        lines.append("NOTE: Indicative only")
    if _maybe(rng, 0.35):
        rng.shuffle(lines)
    return "\n".join(lines)


# -----------------------------
# File helpers
# -----------------------------


def parse_int_range(s: str, *, default: Tuple[int, int] = (2, 6)) -> Tuple[int, int]:
    """Parse 'a-b' or 'n' into (min,max)."""

    s = s.strip()
    if not s:
        return default
    if re.fullmatch(r"\d+", s):
        v = int(s)
        return (v, v)
    m = re.fullmatch(r"(\d+)\s*-\s*(\d+)", s)
    if not m:
        raise ValueError(f"Invalid range '{s}' (expected 'n' or 'a-b')")
    a, b = int(m.group(1)), int(m.group(2))
    if a > b:
        a, b = b, a
    return (a, b)


def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def safe_filename(s: str, max_len: int = 80) -> str:
    s = re.sub(r"[^a-zA-Z0-9._-]+", "_", s).strip("_")
    return s[:max_len] or "file"


def sha1_of_text(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8", errors="ignore")).hexdigest()[:10]


# -----------------------------
# Workbook rendering (layout families)
# -----------------------------


def _apply_default_column_widths(ws, widths: Dict[int, float]) -> None:
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _style_header_cell(cell: Cell) -> None:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = PatternFill("solid", fgColor="DDDDDD")


def _make_placeholder_png(path: Path, *, seed: int, label: str, size: Tuple[int, int] = (160, 160)) -> None:
    """Create a small deterministic placeholder PNG."""

    if not _PIL_AVAILABLE:
        raise RuntimeError("Pillow is not installed")

    rng = random.Random(seed)
    img = PILImage.new("RGB", size, (rng.randint(40, 220), rng.randint(40, 220), rng.randint(40, 220)))
    draw = ImageDraw.Draw(img)
    # Simple geometric marks + label (no font dependencies)
    for _ in range(12):
        x1 = rng.randint(0, size[0] - 1)
        y1 = rng.randint(0, size[1] - 1)
        x2 = rng.randint(x1, size[0] - 1)
        y2 = rng.randint(y1, size[1] - 1)
        draw.rectangle([x1, y1, x2, y2], outline=(0, 0, 0), width=1)
    # Put a short label near bottom
    label = label[:12]
    draw.text((6, size[1] - 18), label, fill=(0, 0, 0))
    img.save(path, format="PNG")


def _try_embed_image(ws, cell_ref: str, image_path: Path) -> Optional[str]:
    """Embed an image if dependencies are available. Returns feature_image filename if embedded."""

    if not (_PIL_AVAILABLE and _OPENPYXL_IMAGE_AVAILABLE):
        return None
    try:
        img = XLImage(str(image_path))
        ws.add_image(img, cell_ref)
        return image_path.name
    except Exception:
        return None


def build_finish_schedule_workbook(
    rng: random.Random,
    schedule_name: str,
    products: List[ProductTruth],
    *,
    with_cover_sheet: bool,
    with_images: bool,
    tmp_dir: Path,
) -> Tuple[Workbook, Dict[str, Any]]:
    """Sample-like layout: SPEC CODE, INDICATIVE IMAGE, ITEM & LOCATION, SPECIFICATIONS, MANUFACTURER/SUPPLIER, NOTES, COST."""

    wb = Workbook()
    # Remove the default sheet to control ordering
    default = wb.active
    wb.remove(default)

    meta: Dict[str, Any] = {"layout": "finish_schedule"}
    if with_cover_sheet:
        cover = wb.create_sheet("Cover Sheet")
        cover["A1"] = "PROJECT: Synthetic Interior Schedule"
        cover["A1"].font = Font(bold=True, size=16)
        cover.merge_cells("A1:F1")
        cover["A3"] = "SCHEDULE NAME"
        cover["B3"] = schedule_name
        cover["A5"] = "REVISION"
        cover["B5"] = rng.choice(["A", "B", "C", "D"]) + str(rng.randint(0, 9))
        cover["A6"] = "DATE"
        cover["B6"] = _dt.date.today().isoformat()
        _apply_default_column_widths(cover, {1: 22, 2: 42, 3: 18, 4: 18, 5: 18, 6: 18})

    ws = wb.create_sheet(rng.choice(["Schedule", "APARTMENTS", "Lighting", "Finishes"]))

    # Header row position + top noise
    header_row = rng.randint(1, 12)
    meta["header_row"] = header_row

    n_cols = 7
    headers = [
        rng.choice(TARGET_HEADERS["doc_code"]),
        rng.choice(TARGET_HEADERS["feature_image"]),
        rng.choice(TARGET_HEADERS["product_description"]),
        rng.choice(TARGET_HEADERS["product_details"]),
        rng.choice(TARGET_HEADERS["brand"]),
        rng.choice(["NOTES/COMMENTS", "NOTES", "COMMENTS"]),
        rng.choice(TARGET_HEADERS["rrp"]),
    ]
    # Trailing spaces sometimes (very realistic)
    if rng.random() < 0.35:
        headers[rng.randrange(len(headers))] = headers[rng.randrange(len(headers))] + " "

    # Add some title/project info rows above header
    if header_row > 1:
        title = f"PROJECT: {rng.choice(['Harbour Apartments', 'Central Office Fitout', 'Coastal Villa', 'Unit Renovation'])}"
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        # Insert a "broken formula" row sometimes
        if header_row > 2 and rng.random() < 0.4:
            ws["A2"] = "=#REF!"
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        # Disclaimer row
        disc_row = min(header_row - 1, 3)
        if disc_row >= 2:
            ws.cell(row=disc_row, column=1).value = (
                "REFER TO DRAWINGS AND SPECIFICATIONS FOR FULL DETAIL.\n"
                "IMAGES AND COSTS ARE INDICATIVE ONLY.\n"
                "VERIFY ON SITE PRIOR TO ORDER."
            )
            ws.cell(row=disc_row, column=1).alignment = Alignment(wrap_text=True)
            ws.merge_cells(start_row=disc_row, start_column=1, end_row=disc_row, end_column=n_cols)

    # Header row
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        _style_header_cell(cell)
        ws.row_dimensions[header_row].height = 24

    # Column widths (image col wider)
    _apply_default_column_widths(ws, {1: 14, 2: 18, 3: 28, 4: 44, 5: 32, 6: 26, 7: 22})

    # Data rows grouped by category with merged category headers
    current_row = header_row + 1
    # Keep stable category order but subset
    by_cat: Dict[str, List[ProductTruth]] = {}
    for p in products:
        cat = p._category or "GENERAL"
        by_cat.setdefault(cat, []).append(p)

    cats = list(by_cat.keys())
    # Shuffle categories sometimes
    if rng.random() < 0.3:
        rng.shuffle(cats)

    for cat in cats:
        # Category merged header
        ws.cell(row=current_row, column=1, value=cat)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=n_cols)
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left")
        current_row += 1

        for prod in by_cat[cat]:
            ws.cell(row=current_row, column=1, value=prod.doc_code)

            # Image cell: embed or blank
            img_cell = ws.cell(row=current_row, column=2)
            if with_images and _PIL_AVAILABLE and _OPENPYXL_IMAGE_AVAILABLE and rng.random() < 0.7:
                # Generate image file
                img_name = f"img_{safe_filename(prod.doc_code)}_{uuid.uuid4().hex[:6]}.png"
                img_path = tmp_dir / img_name
                _make_placeholder_png(img_path, seed=rng.randint(0, 10_000_000), label=prod.doc_code)
                anchored = _try_embed_image(ws, img_cell.coordinate, img_path)
                if anchored:
                    prod.feature_image = anchored
                # Increase row height so images don't overlap too much
                ws.row_dimensions[current_row].height = 90
            else:
                # Put either an image filename or a URL sometimes
                if rng.random() < 0.35:
                    prod.feature_image = prod.feature_image or f"{safe_filename(prod.doc_code)}.jpg"
                    img_cell.value = prod.feature_image
                elif rng.random() < 0.35:
                    prod.feature_image = prod.feature_image or f"https://images.example/{safe_filename(prod.doc_code)}.jpg"
                    img_cell.value = prod.feature_image
                else:
                    img_cell.value = None

            # ITEM & LOCATION
            desc = prod.product_description or ""
            if not desc:
                # Create a location-y description
                room = prod._room or _rng_choice(rng, ROOMS)
                prod._room = room
                if rng.random() < 0.5:
                    desc = f"{prod.product_name or ''} — {room}".strip(" —")
                else:
                    desc = f"ITEM: {prod.product_name or ''}\nLOCATION: {room}".strip()
                prod.product_description = desc
            ws.cell(row=current_row, column=3, value=desc)

            # SPECIFICATIONS
            spec = prod.product_details or spec_block(rng, prod)
            prod.product_details = spec
            c4 = ws.cell(row=current_row, column=4, value=spec)
            c4.alignment = Alignment(wrap_text=True, vertical="top")

            # MANUFACTURER / SUPPLIER
            sup = supplier_block(rng, prod)
            ws.cell(row=current_row, column=5, value=sup).alignment = Alignment(wrap_text=True, vertical="top")

            # NOTES/COMMENTS
            note = rng.choice(
                [
                    "Confirm finish prior to order.",
                    "Install as per manufacturer.",
                    "Allow for wastage.",
                    "Coordinate with services.",
                    "Sample board approval required.",
                    "",
                ]
            )
            ws.cell(row=current_row, column=6, value=note).alignment = Alignment(wrap_text=True, vertical="top")

            # INDICATIVE COST
            cost_repr = price_to_repr(rng, prod.rrp, cat)
            ws.cell(row=current_row, column=7, value=cost_repr)

            current_row += 1

        # Blank row between categories sometimes
        if rng.random() < 0.35:
            current_row += 1

    # Create "used range" bloat sometimes
    if rng.random() < 0.35:
        far_col = rng.randint(25, 60)
        far_row = rng.randint(max(current_row, 120), max(current_row, 400))
        ws.cell(row=far_row, column=far_col, value=" ")
        ws.column_dimensions[get_column_letter(far_col)].width = rng.choice([8, 12, 18])
        meta["used_range_bloat"] = {"row": far_row, "col": far_col}

    return wb, meta


def build_normalized_table_workbook(
    rng: random.Random,
    schedule_name: str,
    products: List[ProductTruth],
    *,
    with_cover_sheet: bool,
    with_images: bool,
    tmp_dir: Path,
) -> Tuple[Workbook, Dict[str, Any]]:
    """Normalized layout with separate columns for most fields."""

    wb = Workbook()
    ws = wb.active
    ws.title = rng.choice(["Schedule", "Product List", "Procurement", "FF&E"])
    meta: Dict[str, Any] = {"layout": "normalized"}

    if with_cover_sheet:
        cover = wb.create_sheet("Cover")
        cover["A1"] = "Cover Sheet"
        cover["A2"] = "Schedule Name"
        cover["B2"] = schedule_name
        cover["A3"] = "Generated"
        cover["B3"] = _dt.datetime.now().isoformat(timespec="seconds")
        ws["A1"] = "='Cover'!B2"  # formula reference like sample2
        ws.merge_cells("A1:N1")
        ws["A1"].font = Font(bold=True, size=14)
        header_row = 3
    else:
        ws["A1"] = schedule_name
        ws.merge_cells("A1:N1")
        ws["A1"].font = Font(bold=True, size=14)
        header_row = 2
    meta["header_row"] = header_row

    # Multi-row header with merged "DIMENSIONS" group sometimes
    use_group_header = rng.random() < 0.55

    columns = [
        ("doc_code", rng.choice(TARGET_HEADERS["doc_code"])),
        ("product_name", rng.choice(TARGET_HEADERS["product_name"])),
        ("brand", rng.choice(TARGET_HEADERS["brand"])),
        ("colour", rng.choice(TARGET_HEADERS["colour"])),
        ("finish", rng.choice(TARGET_HEADERS["finish"])),
        ("material", rng.choice(TARGET_HEADERS["material"])),
        ("width", rng.choice(TARGET_HEADERS["width"])),
        ("length", rng.choice(TARGET_HEADERS["length"])),
        ("height", rng.choice(TARGET_HEADERS["height"])),
        ("qty", rng.choice(TARGET_HEADERS["qty"])),
        ("rrp", rng.choice(TARGET_HEADERS["rrp"])),
        ("product_description", rng.choice(["Room", "Location", "Description"])),
        ("product_details", rng.choice(["Notes", "Details", "Specification"])),
        ("feature_image", rng.choice(["Image", "Image Link", "Photo"])),
    ]

    # Randomly drop some columns (missing values scenario)
    if rng.random() < 0.25:
        drop = rng.choice(["feature_image", "rrp", "finish", "colour"])
        columns = [c for c in columns if c[0] != drop]
        meta["dropped_column"] = drop

    start_col = 1
    if use_group_header:
        # Row with group headers
        ws.cell(row=header_row, column=start_col, value="IDENTIFIERS")
        ws.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=3)
        ws.cell(row=header_row, column=4, value="APPEARANCE")
        ws.merge_cells(start_row=header_row, start_column=4, end_row=header_row, end_column=6)
        # Find positions of dim columns in final list
        dim_cols = [i + 1 for i, (k, _) in enumerate(columns) if k in {"width", "length", "height"}]
        if dim_cols:
            ws.cell(row=header_row, column=min(dim_cols), value="DIMENSIONS")
            ws.merge_cells(start_row=header_row, start_column=min(dim_cols), end_row=header_row, end_column=max(dim_cols))
        # Remaining group
        ws.cell(row=header_row, column=max(1, len(columns) - 3), value="OTHER")
        # Style group header
        for col in range(1, len(columns) + 1):
            c = ws.cell(row=header_row, column=col)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
        header_row2 = header_row + 1
    else:
        header_row2 = header_row

    # Actual header names
    for i, (_, name) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row2, column=i, value=name + (" " if rng.random() < 0.15 else ""))
        _style_header_cell(cell)
    ws.row_dimensions[header_row2].height = 24

    # Data
    row = header_row2 + 1
    for prod in products:
        col_map = {k: idx for idx, (k, _) in enumerate(columns, start=1)}
        if "doc_code" in col_map:
            ws.cell(row=row, column=col_map["doc_code"], value=prod.doc_code)
        if "product_name" in col_map:
            ws.cell(row=row, column=col_map["product_name"], value=prod.product_name)
        if "brand" in col_map:
            ws.cell(row=row, column=col_map["brand"], value=prod.brand)
        if "colour" in col_map:
            ws.cell(row=row, column=col_map["colour"], value=prod.colour)
        if "finish" in col_map:
            ws.cell(row=row, column=col_map["finish"], value=prod.finish)
        if "material" in col_map:
            ws.cell(row=row, column=col_map["material"], value=prod.material)
        if "width" in col_map:
            # Sometimes write as string with units
            ws.cell(row=row, column=col_map["width"], value=(mm_to_repr(rng, prod.width, prod._category or "generic") if rng.random() < 0.6 else prod.width))
        if "length" in col_map:
            ws.cell(row=row, column=col_map["length"], value=(mm_to_repr(rng, prod.length, prod._category or "generic") if rng.random() < 0.6 else prod.length))
        if "height" in col_map:
            ws.cell(row=row, column=col_map["height"], value=(mm_to_repr(rng, prod.height, prod._category or "generic") if rng.random() < 0.6 else prod.height))
        if "qty" in col_map:
            ws.cell(row=row, column=col_map["qty"], value=(prod.qty if rng.random() < 0.8 else str(prod.qty) if prod.qty is not None else None))
        if "rrp" in col_map:
            if rng.random() < 0.6:
                ws.cell(row=row, column=col_map["rrp"], value=float(prod.rrp) if prod.rrp is not None else None)
            else:
                ws.cell(row=row, column=col_map["rrp"], value=price_to_repr(rng, prod.rrp, prod._category or "generic"))
        if "product_description" in col_map:
            ws.cell(row=row, column=col_map["product_description"], value=prod.product_description)
        if "product_details" in col_map:
            ws.cell(row=row, column=col_map["product_details"], value=(prod.product_details or spec_block(rng, prod)))
            ws.cell(row=row, column=col_map["product_details"]).alignment = Alignment(wrap_text=True)
        if "feature_image" in col_map:
            img_cell = ws.cell(row=row, column=col_map["feature_image"])
            if with_images and _PIL_AVAILABLE and _OPENPYXL_IMAGE_AVAILABLE and rng.random() < 0.5:
                img_name = f"img_{safe_filename(prod.doc_code)}_{uuid.uuid4().hex[:6]}.png"
                img_path = tmp_dir / img_name
                _make_placeholder_png(img_path, seed=rng.randint(0, 10_000_000), label=prod.doc_code)
                anchored = _try_embed_image(ws, img_cell.coordinate, img_path)
                if anchored:
                    prod.feature_image = anchored
                    ws.row_dimensions[row].height = 70
            else:
                if rng.random() < 0.5:
                    prod.feature_image = prod.feature_image or f"{safe_filename(prod.doc_code)}.jpg"
                    img_cell.value = prod.feature_image
                else:
                    prod.feature_image = prod.feature_image or f"https://img.example/{safe_filename(prod.doc_code)}.jpg"
                    img_cell.value = prod.feature_image
        row += 1

    # Column widths
    for col in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = rng.choice([12, 14, 18, 22])

    # Hide a column sometimes
    if rng.random() < 0.25:
        col_to_hide = rng.randint(1, len(columns))
        ws.column_dimensions[get_column_letter(col_to_hide)].hidden = True
        meta["hidden_column"] = get_column_letter(col_to_hide)

    return wb, meta


def build_ffe_tracker_workbook(
    rng: random.Random,
    schedule_name: str,
    products: List[ProductTruth],
    *,
    with_images: bool,
    tmp_dir: Path,
) -> Tuple[Workbook, Dict[str, Any]]:
    """FF&E tracker style (often used as Google Sheets)."""

    wb = Workbook()
    ws = wb.active
    ws.title = rng.choice(["FF&E", "Furniture", "Tracker", "Procurement"])  # common
    meta: Dict[str, Any] = {"layout": "ffe_tracker"}

    header_row = rng.randint(1, 10)
    meta["header_row"] = header_row

    # Many FF&E trackers don't have a dedicated code column; we support both.
    include_code_col = rng.random() < 0.6
    meta["include_code_col"] = include_code_col

    total_cols = 10 if include_code_col else 9

    # Top title rows
    if header_row > 1:
        ws["A1"] = f"{schedule_name} (FF&E Tracker)"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        if header_row > 3 and rng.random() < 0.5:
            ws["A2"] = "Prepared by: Synthetic Generator"
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)

    headers = []
    if include_code_col:
        headers.append(rng.choice(TARGET_HEADERS["doc_code"]))
    headers += [
        rng.choice(["Item Name", "Product", "Item"]),
        rng.choice(["Room", "Location", "Area"]),
        rng.choice(["Quantity", "Qty", "Units"]),
        rng.choice(["Dimensions", "Size", "WxDxH"]),
        rng.choice(["Vendor", "Supplier", "Brand"]),
        rng.choice(["Lead Time", "Leadtime", "ETA"]),
        rng.choice(["Cost", "Unit Cost", "Price"]),
        rng.choice(["Image Link", "Photo", "Image"]),
        rng.choice(["Notes", "Comments", "Specification"]),
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h + (" " if rng.random() < 0.2 else ""))
        _style_header_cell(cell)

    row = header_row + 1
    for prod in products:
        room = prod._room or _rng_choice(rng, ROOMS)
        prod._room = room

        dims = None
        if prod.width and prod.length and prod.height and rng.random() < 0.7:
            dims = f"{prod.width}W x {prod.length}D x {prod.height}H mm"
        elif prod.width and prod.length:
            dims = f"{prod.width} x {prod.length} mm"
        else:
            dims = dims_block(rng, prod.width, prod.length, prod.height, prod._category or "generic").replace("\n", "; ")

        lead_time = rng.choice(["2–4 weeks", "6–8 weeks", "8–12 weeks", "In stock", "POA", "TBC"])
        cost_repr = price_to_repr(rng, prod.rrp, prod._category or "generic")

        # Write (column indices depend on whether code col exists)
        col0 = 1
        if include_code_col:
            ws.cell(row=row, column=col0, value=prod.doc_code)
            col0 += 1

        item_name = prod.product_name
        # If there's no code column, embed doc_code into the item name sometimes
        if not include_code_col and prod.doc_code and rng.random() < 0.85:
            item_name = f"{prod.doc_code} - {item_name}" if item_name else prod.doc_code

        ws.cell(row=row, column=col0 + 0, value=item_name)
        ws.cell(row=row, column=col0 + 1, value=room)
        ws.cell(row=row, column=col0 + 2, value=prod.qty)
        ws.cell(row=row, column=col0 + 3, value=dims)
        ws.cell(row=row, column=col0 + 4, value=prod.brand)
        ws.cell(row=row, column=col0 + 5, value=lead_time)
        ws.cell(row=row, column=col0 + 6, value=cost_repr)

        img_cell = ws.cell(row=row, column=col0 + 7)
        if with_images and _PIL_AVAILABLE and _OPENPYXL_IMAGE_AVAILABLE and rng.random() < 0.3:
            img_name = f"img_{safe_filename(prod.doc_code)}_{uuid.uuid4().hex[:6]}.png"
            img_path = tmp_dir / img_name
            _make_placeholder_png(img_path, seed=rng.randint(0, 10_000_000), label=prod.doc_code, size=(120, 120))
            anchored = _try_embed_image(ws, img_cell.coordinate, img_path)
            if anchored:
                prod.feature_image = anchored
                ws.row_dimensions[row].height = 60
        else:
            prod.feature_image = prod.feature_image or f"https://photos.example/{safe_filename(prod.doc_code)}.jpg"
            img_cell.value = prod.feature_image

        note = prod.product_details or rng.choice(
            [
                "Confirm upholstery sample.",
                "Coordinate power requirements.",
                "Mounting height TBC.",
                "Allow for delivery access.",
                "",
            ]
        )
        ws.cell(row=row, column=col0 + 8, value=note)

        # Map to schema-ish truth fields
        prod.product_description = room
        prod.product_details = note

        row += 1

    # Column widths
    widths = {1: 14, 2: 26, 3: 18, 4: 10, 5: 22, 6: 18, 7: 14, 8: 18, 9: 22, 10: 28}
    if not include_code_col:
        widths = {1: 26, 2: 18, 3: 10, 4: 22, 5: 18, 6: 14, 7: 18, 8: 22, 9: 28}
    _apply_default_column_widths(ws, widths)

    # Insert blank rows sometimes
    if rng.random() < 0.35:
        ws.insert_rows(header_row + 1, amount=rng.randint(1, 3))
        meta["blank_rows_after_header"] = True

    return wb, meta


# -----------------------------
# Product generation
# -----------------------------


def _gen_doc_code(rng: random.Random, category: str, idx: int) -> str:
    prefix_map = {
        "FLOORING": "FL",
        "TILES": "TI",
        "WALL FINISHES": "WF",
        "PAINT": "PA",
        "LIGHTING": "L",
        "JOINERY": "J",
        "HARDWARE": "HW",
        "SANITARY": "S",
        "STONE": "ST",
        "GLASS": "G",
        "FURNITURE": "F",
        "APPLIANCES": "A",
    }
    prefix = prefix_map.get(category, "IT")
    return f"{prefix}-{idx:02d}" if rng.random() < 0.75 else f"{prefix}{idx:03d}"


def _gen_dimensions_for_category(rng: random.Random, category: str) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    # Rough, plausible ranges (mm). Include missing occasionally.
    if _maybe(rng, 0.08):
        return None, None, None
    if category in {"TILES", "STONE"}:
        w = rng.choice([300, 450, 600, 900])
        l = rng.choice([300, 450, 600, 900, 1200])
        h = rng.choice([8, 10, 12, 20])
        return w, l, h
    if category == "FLOORING":
        w = rng.choice([120, 180, 220, 300])
        l = rng.choice([1200, 1800, 2200, 2400])
        h = rng.choice([12, 14, 15, 20])
        return w, l, h
    if category == "LIGHTING":
        w = rng.choice([80, 120, 200, 300, 450])
        l = rng.choice([80, 120, 200, 300, 450])
        h = rng.choice([60, 120, 250, 400, 600])
        return w, l, h
    if category == "SANITARY":
        w = rng.choice([400, 500, 600, 750])
        l = rng.choice([350, 450, 520])
        h = rng.choice([120, 150, 180, 900])
        return w, l, h
    if category == "HARDWARE":
        w = rng.choice([20, 32, 60, 80, 120])
        l = rng.choice([20, 32, 60, 80, 120])
        h = rng.choice([10, 20, 30, 40])
        return w, l, h
    if category == "FURNITURE":
        w = rng.choice([450, 600, 800, 1200, 1800])
        l = rng.choice([450, 600, 800, 900])
        h = rng.choice([450, 750, 900])
        return w, l, h
    # Generic
    w = rng.choice([50, 100, 200, 300, 600])
    l = rng.choice([50, 100, 200, 300, 600])
    h = rng.choice([10, 20, 50, 100, 200])
    return w, l, h


def _gen_rrp(rng: random.Random, category: str) -> Optional[Decimal]:
    if _maybe(rng, 0.15):
        return None
    if category in {"FLOORING", "TILES", "STONE"}:
        return Decimal(str(round(_rand_float(rng, 25, 180), 2)))
    if category in {"LIGHTING"}:
        return Decimal(str(round(_rand_float(rng, 60, 1200), 2)))
    if category in {"FURNITURE"}:
        return Decimal(str(round(_rand_float(rng, 120, 4500), 2)))
    if category in {"HARDWARE"}:
        return Decimal(str(round(_rand_float(rng, 12, 260), 2)))
    if category in {"SANITARY"}:
        return Decimal(str(round(_rand_float(rng, 80, 2500), 2)))
    return Decimal(str(round(_rand_float(rng, 20, 800), 2)))


def generate_products(rng: random.Random, n: int) -> List[ProductTruth]:
    products: List[ProductTruth] = []
    # Ensure at least 2 categories
    cats = rng.sample(CATEGORIES, k=min(len(CATEGORIES), max(2, rng.randint(2, 6))))
    for i in range(1, n + 1):
        cat = rng.choice(cats)
        base_type = _rng_choice(rng, PRODUCT_TYPES_BY_CATEGORY.get(cat, ["Product"]))
        name = f"{_rng_choice(rng, ADJECTIVES)} {base_type}" if rng.random() < 0.7 else base_type

        brand = _rng_choice(rng, BRANDS)
        colour = _rng_choice(rng, COLOURS)
        finish = _rng_choice(rng, FINISHES)
        material = _rng_choice(rng, MATERIALS)
        w, l, h = _gen_dimensions_for_category(rng, cat)
        qty = None if _maybe(rng, 0.12) else rng.randint(1, 60)
        rrp = _gen_rrp(rng, cat)

        doc_code = _gen_doc_code(rng, cat, i)

        # Add unicode/special chars sometimes
        if _maybe(rng, 0.08):
            name = name.replace("e", "é") if "e" in name else name + "™"
        if _maybe(rng, 0.05):
            colour = colour + " / " + _rng_choice(rng, ["As shown", "Custom", "TBC"])

        room = _rng_choice(rng, ROOMS)
        desc = rng.choice(
            [
                f"{name} for {room}",
                f"{room}",
                f"Main {room} area",
                f"As per interior elevations ({room})",
            ]
        )

        details = None
        if _maybe(rng, 0.25):
            details = rng.choice(
                [
                    "Install at 2.4m height; dimmable driver included.",
                    "Waterproof membrane required; slip rating R10.",
                    "Allow 10% extra for wastage.",
                    "Supply sample for approval prior to production.",
                ]
            )

        p = ProductTruth(
            doc_code=doc_code,
            product_name=name,
            brand=brand,
            colour=colour,
            finish=finish,
            material=material,
            width=w,
            length=l,
            height=h,
            qty=qty,
            rrp=rrp,
            product_description=desc,
            product_details=details,
            _category=cat,
            _room=room,
        )
        products.append(p)
    return products


# -----------------------------
# Mutation system
# -----------------------------


MutationFn = Callable[[Workbook, random.Random, Optional[ScheduleTruth]], None]


def detect_header_row(ws, *, max_scan_rows: int = 60, max_scan_cols: int = 60) -> Optional[int]:
    """Heuristic: find row with the most known header keywords."""

    best_row: Optional[int] = None
    best_score = 0
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        score = 0
        for c in range(1, min(ws.max_column, max_scan_cols) + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            key = normalize_header_text(v)
            if key in HEADER_LOOKUP:
                score += 1
        if score > best_score:
            best_score = score
            best_row = r
    if best_score >= 3:
        return best_row
    return None


def mutate_insert_noise_rows(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Insert project/title/disclaimer rows above the detected header in random sheets."""

    for ws in wb.worksheets:
        hr = detect_header_row(ws)
        if hr is None:
            continue
        # Insert between 1-4 rows above header
        n = rng.randint(1, 4)
        ws.insert_rows(hr, amount=n)
        # Write title in top inserted row
        title_row = hr
        n_cols = min(12, max(6, ws.max_column))
        ws.cell(row=title_row, column=1, value=f"PROJECT: Mutated Schedule {uuid.uuid4().hex[:6].upper()}")
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=n_cols)
        ws.cell(row=title_row, column=1).font = Font(bold=True, size=14)
        # Disclaimer (merged) in last inserted row
        disc_row = hr + n - 1
        ws.cell(
            row=disc_row,
            column=1,
            value=(
                "IMAGES AND COSTS ARE INDICATIVE ONLY.\n"
                "REFER TO DRAWINGS AND SPECIFICATION.\n"
                "VERIFY ON SITE."
            ),
        )
        ws.cell(row=disc_row, column=1).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=disc_row, start_column=1, end_row=disc_row, end_column=n_cols)


def mutate_rename_headers(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Rename some header cells using synonym pools; add casing/spacing noise."""

    for ws in wb.worksheets:
        hr = detect_header_row(ws)
        if hr is None:
            continue
        max_c = min(ws.max_column, 60)
        for c in range(1, max_c + 1):
            cell = ws.cell(row=hr, column=c)
            v = cell.value
            if not isinstance(v, str):
                continue
            key = normalize_header_text(v)
            target = HEADER_LOOKUP.get(key)
            if not target:
                continue
            if rng.random() < 0.7:
                new_name = rng.choice(TARGET_HEADERS[target])
                # Casing variations
                if rng.random() < 0.25:
                    new_name = new_name.lower().title()
                if rng.random() < 0.15:
                    new_name = new_name.upper()
                # Trailing spaces
                if rng.random() < 0.30:
                    new_name = new_name + " " * rng.randint(1, 2)
                cell.value = new_name


def mutate_hide_rows_cols(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Hide random columns/rows (including sometimes important ones)."""

    for ws in wb.worksheets:
        if ws.max_column < 3:
            continue
        # Hide 0-3 columns
        for _ in range(rng.randint(0, 3)):
            col = rng.randint(1, min(ws.max_column, 30))
            ws.column_dimensions[get_column_letter(col)].hidden = True
        # Hide 0-5 rows (avoid header if detected)
        hr = detect_header_row(ws)
        for _ in range(rng.randint(0, 5)):
            r = rng.randint(1, min(ws.max_row, 200))
            if hr and r == hr:
                continue
            ws.row_dimensions[r].hidden = True


def mutate_add_extra_sheet(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Add a non-data sheet to test multi-sheet scanning."""

    name = rng.choice(["README", "Notes", "Legend", "Revision History", "Calculations"])
    if name in wb.sheetnames:
        name = name + f"_{rng.randint(1, 9)}"
    ws = wb.create_sheet(name)
    ws["A1"] = "This sheet is not part of the schedule."
    ws["A2"] = "It exists to test sheet-selection heuristics."
    ws["A4"] = "Key"
    ws["B4"] = "Value"
    for i in range(5, 15):
        ws[f"A{i}"] = rng.choice(["Revision", "Client", "Prepared by", "Approved", "Date"])
        ws[f"B{i}"] = rng.choice(["TBC", "N/A", "See cover sheet", _dt.date.today().isoformat()])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 36


def mutate_add_cover_sheet_and_formula(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Add a cover sheet and reference schedule name via formula in a data sheet."""

    if "Cover Sheet" in wb.sheetnames:
        return
    cover = wb.create_sheet("Cover Sheet", 0)
    cover["A1"] = "PROJECT: Mutated Project"
    cover["A1"].font = Font(bold=True, size=16)
    cover.merge_cells("A1:F1")
    cover["A3"] = "SCHEDULE NAME"
    cover["B3"] = truth.schedule_name if truth else rng.choice(["Lighting Schedule", "Finishes Schedule", "FF&E"]) 
    cover["A5"] = "DATE"
    cover["B5"] = _dt.date.today().isoformat()

    # Find a likely data sheet (first non-cover)
    target_ws = None
    for ws in wb.worksheets:
        if ws.title != "Cover Sheet":
            target_ws = ws
            break
    if not target_ws:
        return
    # Put formula in top-left
    target_ws.insert_rows(1)
    target_ws["A1"] = "='Cover Sheet'!B3"
    target_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(12, max(6, target_ws.max_column)))
    target_ws["A1"].font = Font(bold=True, size=14)


def mutate_merge_random_title_cells(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Merge some cells in top rows (title/project info blocks)."""

    for ws in wb.worksheets:
        if ws.max_column < 3:
            continue
        top_rows = min(5, ws.max_row)
        if top_rows < 1:
            continue
        r = rng.randint(1, top_rows)
        c1 = 1
        c2 = rng.randint(min(3, ws.max_column), min(ws.max_column, 12))
        try:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            ws.cell(row=r, column=c1).alignment = Alignment(wrap_text=True)
        except Exception:
            pass


def mutate_insert_blank_rows(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Insert blank rows between sections/data."""

    for ws in wb.worksheets:
        hr = detect_header_row(ws)
        if hr is None:
            continue
        # Insert blank rows somewhere after header
        if ws.max_row <= hr + 5:
            continue
        for _ in range(rng.randint(1, 4)):
            pos = rng.randint(hr + 2, min(ws.max_row, hr + 60))
            ws.insert_rows(pos)


def mutate_repeat_header_mid_sheet(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Repeat the header row mid-sheet (common in printed sections)."""

    for ws in wb.worksheets:
        hr = detect_header_row(ws)
        if hr is None:
            continue
        if ws.max_row < hr + 30:
            continue
        # Copy header values
        header_vals = [ws.cell(row=hr, column=c).value for c in range(1, min(ws.max_column, 30) + 1)]
        insert_at = rng.randint(hr + 10, min(ws.max_row, hr + 80))
        ws.insert_rows(insert_at)
        for c, v in enumerate(header_vals, start=1):
            ws.cell(row=insert_at, column=c, value=v)
            ws.cell(row=insert_at, column=c).font = Font(bold=True)
            ws.cell(row=insert_at, column=c).fill = PatternFill("solid", fgColor="EEEEEE")


def mutate_add_category_rows(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Insert merged category headers spanning the full width."""

    for ws in wb.worksheets:
        hr = detect_header_row(ws)
        if hr is None:
            continue
        width = min(ws.max_column, 20)
        if width < 4:
            continue
        # Insert 1-4 category rows
        for _ in range(rng.randint(1, 4)):
            pos = rng.randint(hr + 1, min(ws.max_row, hr + 120))
            ws.insert_rows(pos)
            cat = rng.choice(CATEGORIES)
            ws.cell(row=pos, column=1, value=cat)
            try:
                ws.merge_cells(start_row=pos, start_column=1, end_row=pos, end_column=width)
            except Exception:
                pass
            ws.cell(row=pos, column=1).font = Font(bold=True)


def mutate_change_some_price_formats(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Find likely price columns and perturb formatting (numbers -> strings, strings -> numbers)."""

    for ws in wb.worksheets:
        hr = detect_header_row(ws)
        if hr is None:
            continue
        # Identify price-like columns by header text
        price_cols: List[int] = []
        for c in range(1, min(ws.max_column, 60) + 1):
            v = ws.cell(row=hr, column=c).value
            if isinstance(v, str) and "COST" in normalize_header_text(v) or (isinstance(v, str) and normalize_header_text(v) in {"RRP", "PRICE"}):
                price_cols.append(c)
        if not price_cols:
            continue
        col = rng.choice(price_cols)
        # perturb some rows
        for r in range(hr + 1, min(ws.max_row, hr + 80) + 1):
            cell = ws.cell(row=r, column=col)
            if cell.value is None:
                continue
            if isinstance(cell.value, (int, float)) and rng.random() < 0.7:
                cell.value = price_to_repr(rng, Decimal(str(float(cell.value))), "generic")
            elif isinstance(cell.value, str) and rng.random() < 0.25:
                m = re.search(r"(\d+(?:\.\d+)?)", cell.value)
                if m:
                    try:
                        cell.value = float(m.group(1))
                    except Exception:
                        pass


def mutate_expand_used_range(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Expand 'used range' artificially by writing a whitespace cell far away."""

    for ws in wb.worksheets:
        far_row = rng.randint(max(200, ws.max_row), max(800, ws.max_row + 300))
        far_col = rng.randint(max(30, ws.max_column), max(80, ws.max_column + 20))
        ws.cell(row=far_row, column=far_col, value=" ")
        ws.column_dimensions[get_column_letter(far_col)].width = rng.choice([8, 12, 16])


def mutate_insert_excel_error_cell(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Insert a cell containing an Excel-like error / broken formula near the top."""

    ws = rng.choice(wb.worksheets)
    r = rng.randint(1, min(10, ws.max_row))
    c = rng.randint(1, min(10, ws.max_column))
    ws.cell(row=r, column=c, value=rng.choice(["=#REF!", "#N/A", "#VALUE!", "=1/0"]))


def mutate_add_totals_row(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Append a totals row with formulas (common in real schedules).

    We don't rely on formulas being evaluated; the goal is to include them so
    parsers don't choke on formula cells.
    """

    for ws in wb.worksheets:
        hr = detect_header_row(ws)
        if hr is None:
            continue

        # Identify qty/price columns by header mapping
        qty_cols: List[int] = []
        price_cols: List[int] = []
        for c in range(1, min(ws.max_column, 60) + 1):
            hv = ws.cell(row=hr, column=c).value
            if not isinstance(hv, str):
                continue
            key = normalize_header_text(hv)
            target = HEADER_LOOKUP.get(key)
            if target == "qty":
                qty_cols.append(c)
            if target == "rrp":
                price_cols.append(c)

        if not (qty_cols or price_cols):
            continue

        # Find a plausible data region end (last non-empty row under header)
        data_start = hr + 1
        data_end = None
        for r in range(min(ws.max_row, hr + 600), data_start - 1, -1):
            row_has_data = False
            for c in range(1, min(ws.max_column, 25) + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip() == "":
                    continue
                if v is not None:
                    row_has_data = True
                    break
            if row_has_data:
                data_end = r
                break
        if data_end is None or data_end <= data_start:
            continue

        total_row = data_end + 1
        # Label
        ws.cell(row=total_row, column=1, value=rng.choice(["TOTAL", "TOTALS", "SUBTOTAL"])).font = Font(bold=True)

        def _sum_formula(col: int) -> str:
            letter = get_column_letter(col)
            return f"=SUM({letter}{data_start}:{letter}{data_end})"

        for col in qty_cols:
            ws.cell(row=total_row, column=col, value=_sum_formula(col)).font = Font(bold=True)
        for col in price_cols:
            ws.cell(row=total_row, column=col, value=_sum_formula(col)).font = Font(bold=True)

        # Make it visually distinct
        for c in range(1, min(ws.max_column, 25) + 1):
            ws.cell(row=total_row, column=c).fill = PatternFill("solid", fgColor="F2F2F2")


def mutate_add_long_text_cell(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Inject a very long, messy text blob into a likely 'notes/spec' column."""

    ws = rng.choice(wb.worksheets)
    hr = detect_header_row(ws)
    if hr is None:
        return

    # Find a likely notes/spec/details column
    candidate_cols: List[int] = []
    for c in range(1, min(ws.max_column, 60) + 1):
        hv = ws.cell(row=hr, column=c).value
        if not isinstance(hv, str):
            continue
        tgt = HEADER_LOOKUP.get(normalize_header_text(hv))
        if tgt in {"product_details", "product_description"}:
            candidate_cols.append(c)
        # Some sheets label notes without mapping
        if "NOTE" in normalize_header_text(hv) or "COMMENT" in normalize_header_text(hv):
            candidate_cols.append(c)

    if not candidate_cols:
        candidate_cols = [rng.randint(1, min(ws.max_column, 12))]

    col = rng.choice(candidate_cols)
    # pick a data-ish row
    r = rng.randint(hr + 1, min(ws.max_row, hr + 80))

    blob = (
        "IMPORTANT NOTES / EDGE CASE TEST\n"
        "• This cell intentionally contains very long text, line breaks, unicode, and punctuation.\n"
        "• Characters: café, façade, Łódź, ✓, ™, —, …, Ø, å.\n"
        "• Mixed delimiters: KEY:VALUE | KEY - VALUE | KEY=VALUE\n\n"
        "Lorem ipsum dolor sit amet, consectetur adipiscing elit. "
        "Sed do eiusmod tempor incididunt ut labore et dolore magna aliqua. "
    )
    # Expand to ~1500-2500 chars
    blob = blob + (" Lorem ipsum — " * rng.randint(80, 140))
    ws.cell(row=r, column=col, value=blob).alignment = Alignment(wrap_text=True, vertical="top")


def mutate_add_sales_schedule_variant(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Create a second schedule-like sheet (e.g., 'Sales Schedule ') with hidden columns.

    This mimics workbooks that contain multiple similar schedules with slightly different
    visibility settings.
    """

    # Avoid duplicating if already present
    if any("SALES" in ws.title.upper() for ws in wb.worksheets):
        return

    # Pick a source sheet that looks like data
    src = None
    for ws in wb.worksheets:
        if detect_header_row(ws) is not None:
            src = ws
            break
    if src is None:
        return

    title = "Sales Schedule "  # note trailing space (seen in sample2)
    if title in wb.sheetnames:
        title = f"Sales Schedule {rng.randint(1,9)} "

    dst = wb.create_sheet(title)

    # Copy a bounded region (values only) to avoid heavy files
    max_r = min(src.max_row, 300)
    max_c = min(src.max_column, 25)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            dst.cell(row=r, column=c, value=src.cell(row=r, column=c).value)

    # Copy merged ranges within the copied bounds (best-effort)
    for rng_cellrange in getattr(src.merged_cells, "ranges", []):
        try:
            if rng_cellrange.max_row <= max_r and rng_cellrange.max_col <= max_c:
                dst.merge_cells(str(rng_cellrange))
        except Exception:
            continue

    # Hide some columns
    for _ in range(rng.randint(1, 5)):
        col = rng.randint(1, max_c)
        dst.column_dimensions[get_column_letter(col)].hidden = True

    # Hide a few rows as well
    for _ in range(rng.randint(0, 3)):
        row = rng.randint(1, min(max_r, 60))
        dst.row_dimensions[row].hidden = True


def mutate_sheetname_trailing_space(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Add trailing whitespace to a random sheet name (Excel allows it)."""

    ws = rng.choice(wb.worksheets)
    if ws.title.endswith(" "):
        return
    # Excel limits sheet names to 31 chars
    if len(ws.title) >= 31:
        return
    try:
        ws.title = ws.title + " "
    except Exception:
        return


def mutate_swap_two_columns(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Swap two random columns' *values* (simulates manual column re-ordering)."""

    ws = rng.choice(wb.worksheets)
    if ws.max_column < 4 or ws.max_row < 8:
        return

    c1 = rng.randint(1, min(ws.max_column, 15))
    c2 = rng.randint(1, min(ws.max_column, 15))
    if c1 == c2:
        return

    max_r = min(ws.max_row, 250)
    try:
        from openpyxl.cell.cell import MergedCell  # type: ignore
    except Exception:
        MergedCell = ()  # type: ignore

    for r in range(1, max_r + 1):
        cell1 = ws.cell(row=r, column=c1)
        cell2 = ws.cell(row=r, column=c2)
        if MergedCell and (isinstance(cell1, MergedCell) or isinstance(cell2, MergedCell)):
            continue
        v1, v2 = cell1.value, cell2.value
        try:
            cell1.value, cell2.value = v2, v1
        except Exception:
            continue


def mutate_insert_blank_columns(wb: Workbook, rng: random.Random, truth: Optional[ScheduleTruth]) -> None:
    """Insert blank columns near the left side (common when people add notes/images)."""

    ws = rng.choice(wb.worksheets)
    if ws.max_column < 3:
        return
    pos = rng.randint(1, min(6, ws.max_column))
    amount = rng.randint(1, 2)
    try:
        ws.insert_cols(pos, amount=amount)
    except Exception:
        return


MUTATIONS: List[Tuple[str, MutationFn]] = [
    ("insert_noise_rows", mutate_insert_noise_rows),
    ("rename_headers", mutate_rename_headers),
    ("hide_rows_cols", mutate_hide_rows_cols),
    ("add_extra_sheet", mutate_add_extra_sheet),
    ("add_cover_sheet_formula", mutate_add_cover_sheet_and_formula),
    ("add_sales_schedule_variant", mutate_add_sales_schedule_variant),
    ("sheetname_trailing_space", mutate_sheetname_trailing_space),
    ("merge_title_cells", mutate_merge_random_title_cells),
    ("insert_blank_rows", mutate_insert_blank_rows),
    ("insert_blank_columns", mutate_insert_blank_columns),
    ("swap_two_columns", mutate_swap_two_columns),
    ("repeat_header_mid_sheet", mutate_repeat_header_mid_sheet),
    ("add_category_rows", mutate_add_category_rows),
    ("change_price_formats", mutate_change_some_price_formats),
    ("add_totals_row", mutate_add_totals_row),
    ("add_long_text_cell", mutate_add_long_text_cell),
    ("expand_used_range", mutate_expand_used_range),
    ("insert_excel_error", mutate_insert_excel_error_cell),
]


def apply_mutations(
    wb: Workbook,
    rng: random.Random,
    truth: Optional[ScheduleTruth],
    *,
    min_mutations: int,
    max_mutations: int,
) -> List[str]:
    n = rng.randint(min_mutations, max_mutations)
    chosen = rng.sample(MUTATIONS, k=min(n, len(MUTATIONS)))
    applied: List[str] = []
    for name, fn in chosen:
        try:
            fn(wb, rng, truth)
            applied.append(name)
        except Exception as e:
            # Log mutation failures for debugging while allowing graceful degradation.
            # The point is to create many varied files, not to stop on one edge case.
            logger.warning(f"Mutation '{name}' failed: {type(e).__name__}: {e}")
            continue
    return applied


# -----------------------------
# Serialization
# -----------------------------


def truth_to_dict(truth: ScheduleTruth) -> Dict[str, Any]:
    def _prod(p: ProductTruth) -> Dict[str, Any]:
        d = dataclasses.asdict(p)
        # Remove internal fields
        d.pop("_category", None)
        d.pop("_room", None)
        # Decimal -> float for JSON
        if d.get("rrp") is not None:
            d["rrp"] = float(d["rrp"])  # type: ignore[assignment]
        return d

    return {
        "schedule_name": truth.schedule_name,
        "layout_family": truth.layout_family,
        "seed": truth.seed,
        "mutations": truth.mutations,
        "notes": truth.notes,
        "products": [_prod(p) for p in truth.products],
    }


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def append_manifest(path: Path, record: Dict[str, Any]) -> None:
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")


# -----------------------------
# Generation orchestration
# -----------------------------


LAYOUT_BUILDERS: List[Tuple[str, Callable[..., Tuple[Workbook, Dict[str, Any]]]]] = [
    ("finish_schedule", build_finish_schedule_workbook),
    ("normalized", build_normalized_table_workbook),
    ("ffe_tracker", build_ffe_tracker_workbook),
]


def generate_one_workbook(
    *,
    seed: int,
    out_path: Path,
    truth_path: Path,
    manifest_path: Path,
    with_images: bool,
    min_products: int,
    max_products: int,
    mutation_range: Tuple[int, int],
    allow_cover_sheet: bool,
    layout_choice: Optional[str] = None,
) -> None:
    rng = random.Random(seed)

    schedule_name = rng.choice(
        [
            "Lighting Schedule",
            "Finishes Schedule",
            "Apartment Schedule",
            "Sales Schedule",
            "FF&E Schedule",
            "Interior Schedule",
        ]
    )
    n_products = rng.randint(min_products, max_products)
    products = generate_products(rng, n_products)

    # Choose layout
    if layout_choice:
        builder = next((b for name, b in LAYOUT_BUILDERS if name == layout_choice), None)
        if not builder:
            raise ValueError(f"Unknown layout '{layout_choice}'")
        layout_family = layout_choice
    else:
        layout_family, builder = rng.choice(LAYOUT_BUILDERS)

    tmp_dir = out_path.parent / "_tmp_images"
    ensure_dir(tmp_dir)

    # Some layouts use cover sheets
    with_cover = allow_cover_sheet and rng.random() < 0.55 and layout_family in {"finish_schedule", "normalized"}

    if layout_family == "ffe_tracker":
        wb, meta = builder(rng, schedule_name, products, with_images=with_images, tmp_dir=tmp_dir)  # type: ignore[arg-type]
    else:
        wb, meta = builder(
            rng,
            schedule_name,
            products,
            with_cover_sheet=with_cover,
            with_images=with_images,
            tmp_dir=tmp_dir,
        )  # type: ignore[arg-type]

    truth = ScheduleTruth(
        schedule_name=schedule_name,
        products=products,
        layout_family=layout_family,
        seed=seed,
        mutations=[],
        notes=meta,
    )

    # Apply mutations
    applied = apply_mutations(wb, rng, truth, min_mutations=mutation_range[0], max_mutations=mutation_range[1])
    truth.mutations = applied

    # Save
    wb.save(out_path)
    write_json(truth_path, truth_to_dict(truth))

    # Best-effort cleanup: remove placeholder images we created.
    # The images are embedded inside the workbook; the PNGs are just temporary inputs.
    # We keep the folder around to avoid race-y mkdir/remove cycles.
    try:
        for p in tmp_dir.glob("img_*.png"):
            p.unlink()
    except Exception:
        pass

    # Record manifest
    append_manifest(
        manifest_path,
        {
            "type": "generated",
            "xlsx": str(out_path),
            "truth": str(truth_path),
            "seed": seed,
            "layout_family": layout_family,
            "mutations": applied,
        },
    )


def mutate_one_sample(
    *,
    sample_path: Path,
    seed: int,
    out_path: Path,
    meta_path: Path,
    manifest_path: Path,
    mutation_range: Tuple[int, int],
) -> None:
    rng = random.Random(seed)

    wb = load_workbook(sample_path)
    # We don't have reliable ground-truth for arbitrary samples; we emit metadata only.
    applied = apply_mutations(wb, rng, truth=None, min_mutations=mutation_range[0], max_mutations=mutation_range[1])
    wb.save(out_path)

    meta = {
        "type": "mutated",
        "source": str(sample_path),
        "seed": seed,
        "mutations": applied,
        "notes": {
            "sheetnames": wb.sheetnames,
        },
    }
    write_json(meta_path, meta)

    append_manifest(
        manifest_path,
        {
            "type": "mutated",
            "source": str(sample_path),
            "xlsx": str(out_path),
            "meta": str(meta_path),
            "seed": seed,
            "mutations": applied,
        },
    )


# -----------------------------
# CLI
# -----------------------------


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Generate and/or mutate Excel interior schedule files for robust parser testing.",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    p.add_argument(
        "--mode",
        choices=["generate", "mutate", "both"],
        default="generate",
        help="Whether to generate new workbooks, mutate provided samples, or both.",
    )
    p.add_argument(
        "--samples_dir",
        type=str,
        default=None,
        help="Directory containing .xlsx samples/templates to mutate (required for mode=mutate or both).",
    )
    p.add_argument("--output_dir", type=str, required=True, help="Output directory.")
    p.add_argument("--seed", type=int, default=12345, help="Base RNG seed (reproducible).")
    p.add_argument(
        "--mutations",
        type=str,
        default="2-6",
        help="How many mutations to apply per output file (e.g. '3' or '2-6').",
    )
    p.add_argument("--num_generated", type=int, default=20, help="Number of brand-new synthetic workbooks to generate.")
    p.add_argument(
        "--mutants_per_sample",
        type=int,
        default=3,
        help="How many mutated variants to produce per sample workbook.",
    )
    p.add_argument("--min_products", type=int, default=15, help="Min products per generated workbook.")
    p.add_argument("--max_products", type=int, default=45, help="Max products per generated workbook.")
    p.add_argument(
        "--with_images",
        action="store_true",
        help="Attempt to embed placeholder images in some rows (requires Pillow).",
    )
    p.add_argument(
        "--no_cover_sheet",
        action="store_true",
        help="Disable generating cover sheets / cross-sheet formula references.",
    )
    p.add_argument(
        "--layout",
        type=str,
        default=None,
        choices=[None, "finish_schedule", "normalized", "ffe_tracker"],
        help="Force a specific layout family for generated workbooks.",
    )
    return p


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = build_arg_parser().parse_args(argv)
    out_dir = Path(args.output_dir)
    ensure_dir(out_dir)

    generated_dir = out_dir / "generated"
    mutated_dir = out_dir / "mutated"
    ensure_dir(generated_dir)
    ensure_dir(mutated_dir)

    manifest_path = out_dir / "manifest.jsonl"
    if manifest_path.exists():
        manifest_path.unlink()

    mutation_range = parse_int_range(args.mutations)
    allow_cover = not args.no_cover_sheet

    # Generate
    if args.mode in {"generate", "both"}:
        for i in range(args.num_generated):
            seed = args.seed + i
            base_name = f"gen_{i:04d}_{args.layout or 'mix'}_{seed}"
            xlsx_path = generated_dir / f"{base_name}.xlsx"
            truth_path = generated_dir / f"{base_name}.truth.json"
            generate_one_workbook(
                seed=seed,
                out_path=xlsx_path,
                truth_path=truth_path,
                manifest_path=manifest_path,
                with_images=bool(args.with_images),
                min_products=args.min_products,
                max_products=args.max_products,
                mutation_range=mutation_range,
                allow_cover_sheet=allow_cover,
                layout_choice=args.layout,
            )

    # Mutate
    if args.mode in {"mutate", "both"}:
        if not args.samples_dir:
            print("ERROR: --samples_dir is required for mode=mutate or both", file=sys.stderr)
            return 2
        samples_dir = Path(args.samples_dir)
        if not samples_dir.exists():
            print(f"ERROR: samples_dir does not exist: {samples_dir}", file=sys.stderr)
            return 2

        sample_files = sorted([p for p in samples_dir.glob("*.xlsx") if p.is_file()])
        if not sample_files:
            print(f"ERROR: No .xlsx files found in {samples_dir}", file=sys.stderr)
            return 2

        mutant_idx = 0
        for sp in sample_files:
            for j in range(args.mutants_per_sample):
                seed = args.seed + 10_000 + mutant_idx
                mutant_idx += 1
                base = safe_filename(sp.stem)
                out_name = f"mut_{base}_{j:02d}_{seed}"
                xlsx_path = mutated_dir / f"{out_name}.xlsx"
                meta_path = mutated_dir / f"{out_name}.meta.json"
                mutate_one_sample(
                    sample_path=sp,
                    seed=seed,
                    out_path=xlsx_path,
                    meta_path=meta_path,
                    manifest_path=manifest_path,
                    mutation_range=mutation_range,
                )

    print(f"Done. Output written to: {out_dir}")
    print(f"Manifest: {manifest_path}")
    if args.with_images and not _PIL_AVAILABLE:
        print("NOTE: Pillow not available; images were not embedded.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
